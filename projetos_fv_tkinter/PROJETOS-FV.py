from tkinter import *
import tkinter as tk
from tkinter import ttk

import xlsxwriter

import openpyxl as op

import time

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import math

import pyautogui


class Application:
    def __init__(self, master=None):

        # TELA INICIAL

        self.fontePadrao = ("Arial", "10")

        #self.titulo = Label(root, text="MOTORMAC", foreground='white', background='black')
        #self.titulo["font"] = ("Arial", "30", "bold")
        #self.titulo.place(x=275, y=20)

        self.introd = Label(root,text="""Olá, seja bem vindo ao software de projetos fotovoltaicos!\nSelecione abaixo a opção desejada.""",foreground='white', background='gray18')
        self.introd["font"] = ("Arial", "15", "bold")
        self.introd.place(x=90, y=320)

        self.janela_cadastro = Button(root, foreground='darkorange2', background='white')
        self.janela_cadastro["text"] = "Cadastro Equipamentos"
        self.janela_cadastro["font"] = ("Calibri", "10", "bold")
        self.janela_cadastro["height"] = 2
        self.janela_cadastro["width"] = 30
        self.janela_cadastro["command"] = self.Janela_cadastro
        self.janela_cadastro.place(x=46, y=420)

        self.janela_cliente = Button(root, foreground='darkorange2', background='white')
        self.janela_cliente["text"] = "Cadastro Monitoramento Remoto"
        self.janela_cliente["font"] = ("Calibri", "10", "bold")
        self.janela_cliente["height"] = 2
        self.janela_cliente["width"] = 30
        self.janela_cliente["command"] = self.Janela_monit_remoto
        self.janela_cliente.place(x=265, y=420)

        self.janela_string = Button(root, foreground='darkorange2', background='white')
        self.janela_string["text"] = "Configuração de String"
        self.janela_string["font"] = ("Calibri", "10", "bold")
        self.janela_string["height"] = 2
        self.janela_string["width"] = 30
        self.janela_string["command"] = self.Janela_string
        self.janela_string.place(x=484, y=420)

        self.janelaCC = Button(root, foreground='darkorange2', background='white')
        self.janelaCC["text"] = "Dimensionamento CC"
        self.janelaCC["font"] = ("Calibri", "10", "bold")
        self.janelaCC["height"] = 2
        self.janelaCC["width"] = 30
        self.janelaCC["command"] = self.JanelaCC
        self.janelaCC.place(x=46, y=460)

        self.janelaCA = Button(root, foreground='darkorange2', background='white')
        self.janelaCA["text"] = "Dimensionamento CA"
        self.janelaCA["font"] = ("Calibri", "10", "bold")
        self.janelaCA["height"] = 2
        self.janelaCA["width"] = 30
        self.janelaCA["command"] = self.JanelaCA
        self.janelaCA.place(x=265, y=460)

        self.janela_localizacao = Button(root, foreground='darkorange2', background='white')
        self.janela_localizacao["text"] = "Localização"
        self.janela_localizacao["font"] = ("Calibri", "10", "bold")
        self.janela_localizacao["height"] = 2
        self.janela_localizacao["width"] = 30
        self.janela_localizacao["command"] = self.Janela_localizacao
        self.janela_localizacao.place(x=484, y=460)

        self.janela_unifilar = Button(root, foreground='darkorange2', background='white')
        self.janela_unifilar["text"] = "Automação Unifilar"
        self.janela_unifilar["font"] = ("Calibri", "10", "bold")
        self.janela_unifilar["height"] = 2
        self.janela_unifilar["width"] = 30
        self.janela_unifilar["command"] = self.Janela_automacao
        self.janela_unifilar.place(x=46, y=500)

        self.janela_projecao = Button(root, foreground='darkorange2', background='white')
        self.janela_projecao["text"] = "Projeção plano horizontal"
        self.janela_projecao["font"] = ("Calibri", "10", "bold")
        self.janela_projecao["height"] = 2
        self.janela_projecao["width"] = 30
        self.janela_projecao["command"] = self.Janela_projecao
        self.janela_projecao.place(x=265, y=500)




    def JanelaCA(self):
        self.JanelaCA = Toplevel()
        self.JanelaCA.geometry("760x510")
        self.JanelaCA.minsize(760, 510)
        self.JanelaCA.maxsize(760, 510)
        self.JanelaCA.configure(bg='gray18')
        self.JanelaCA.title("Dimensionamento CA")
        self.JanelaCA.iconbitmap('baixados.ico')

     # CAPACIDADE DE CONDUÇÃO DE CORRENTE

        self.fontePadrao = ("Arial", "10")

        self.titulo = Label(self.JanelaCA, text="Capacidade de condução de corrente", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=40, y=5)

        self.nome2Label = Label(self.JanelaCA, text="Corrente de projeto:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome2Label.place(x=30, y=35)

        self.nome2 = Entry(self.JanelaCA)
        self.nome2["width"] = 10
        self.nome2["font"] = self.fontePadrao
        self.nome2.place(x=240, y=35)

        self.nome3Label = Label(self.JanelaCA, text="Fator de correção de agrupamento:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome3Label.place(x=30, y=65)

        self.nome3 = Entry(self.JanelaCA)
        self.nome3["width"] = 10
        self.nome3["font"] = self.fontePadrao
        self.nome3.place(x=240, y=65)

        self.nome4Label = Label(self.JanelaCA, text="Fator de correção de temperatura:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome4Label.place(x=30, y=95)

        self.nome4 = Entry(self.JanelaCA)
        self.nome4["width"] = 10
        self.nome4["font"] = self.fontePadrao
        self.nome4.place(x=240, y=95)

        self.nome5Label = Label(self.JanelaCA, text="Método:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome5Label.place(x=30, y=125)

        metodo = ['Selecione', 'A1', 'A2', 'B1', 'B2', 'C', 'D']

        self.variable2 = tk.StringVar(self.JanelaCA)
        self.variable2.set(metodo[0])

        self.nome5 = ttk.Combobox(self.JanelaCA, textvariable=self.variable2)
        self.nome5["width"] = 9
        self.nome5.place(x=240, y=125)

        self.nome5['values'] = [metodo[m] for m in range(0, 7)]

        #self.nome5 = Entry(root)
        #self.nome5["width"] = 10
        #self.nome5["font"] = self.fontePadrao
        #self.nome5.place(x=85, y=125)

        self.nome6Label = Label(self.JanelaCA, text="N° de condutores carregados:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome6Label.place(x=30, y=155)

        self.nome6 = Entry(self.JanelaCA)
        self.nome6["width"] = 10
        self.nome6["font"] = self.fontePadrao
        self.nome6.place(x=240, y=155)

        self.nome7Label = Label(self.JanelaCA, text="Isolação:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome7Label.place(x=30, y=185)

        isolacao = ['Selecione', 'PVC', 'EPR/XLPE']

        self.variable3 = tk.StringVar(self.JanelaCA)
        self.variable3.set(isolacao[0])

        self.nome7 = ttk.Combobox(self.JanelaCA, textvariable=self.variable3)
        self.nome7["width"] = 9
        self.nome7.place(x=240, y=185)

        self.nome7['values'] = [isolacao[m] for m in range(0, 3)]

        #self.nome7 = Entry(root)
        #self.nome7["width"] = 10
        #self.nome7["font"] = self.fontePadrao
        #self.nome7.place(x=90, y=185)

        self.calcular = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.calcular["text"] = "Calcular"
        self.calcular["font"] = ("Calibri", "8", "bold")
        self.calcular["width"] = 12
        self.calcular["command"] = self.Calcular
        self.calcular.place(x=100, y=215)

        self.mensagem = Label(self.JanelaCA, text="", font=self.fontePadrao, fg='deep sky blue', bg='gray18')
        self.mensagem.place(x=30, y=245)


        # QUEDA DE TENSÃO

        self.titulo = Label(self.JanelaCA, text="Queda de tensão", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=470,y=5)

        self.nome8Label = Label(self.JanelaCA, text="Tensão:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome8Label.place(x=400,y=35)

        self.nome8 = Entry(self.JanelaCA)
        self.nome8["width"] = 10
        self.nome8["font"] = self.fontePadrao
        self.nome8.place(x=595,y=35)

        self.nome9Label = Label(self.JanelaCA, text="Queda de tensão tabelada:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome9Label.place(x=400,y=65)

        self.nome9 = Entry(self.JanelaCA)
        self.nome9["width"] = 10
        self.nome9["font"] = self.fontePadrao
        self.nome9.place(x=595,y=65)

        self.nome10Label = Label(self.JanelaCA, text="Comprimento do circuito em km:",font=self.fontePadrao, fg='white', bg='gray18')
        self.nome10Label.place(x=400,y=95)

        self.nome10 = Entry(self.JanelaCA)
        self.nome10["width"] = 10
        self.nome10["font"] = self.fontePadrao
        self.nome10.place(x=595,y=95)

        self.calcular1 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.calcular1["text"] = "Calcular"
        self.calcular1["font"] = ("Calibri", "8", "bold")
        self.calcular1["width"] = 12
        self.calcular1["command"] = self.Calcular1
        self.calcular1.place(x=450,y=125)

        self.mensagem1 = Label(self.JanelaCA, text="", font=self.fontePadrao, fg='deep sky blue', bg='gray18')
        self.mensagem1.place(x=400,y=155)


        # SOBRECARGA

        self.titulo = Label(self.JanelaCA, text="Sobrecarga", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=95, y=265)

        self.nome11Label = Label(self.JanelaCA, text="Capacidade de condução de corrente:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome11Label.place(x=30, y=295)

        self.nome11 = Entry(self.JanelaCA)
        self.nome11["width"] = 10
        self.nome11["font"] = self.fontePadrao
        self.nome11.place(x=255, y=295)

        self.calcular2 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.calcular2["text"] = "Calcular"
        self.calcular2["font"] = ("Calibri", "8", "bold")
        self.calcular2["width"] = 12
        self.calcular2["command"] = self.Calcular2
        self.calcular2.place(x=100, y=325)

        self.mensagem2 = Label(self.JanelaCA, text="", font=self.fontePadrao, fg='deep sky blue', bg='gray18')
        self.mensagem2.place(x=30, y=355)

        self.nome14Label = Label(self.JanelaCA,text="Confirme o disjuntor: ",font=("Arial", "10", "bold", "italic"), fg='white', bg='gray18')
        self.nome14Label.place(x=30, y=400)

        disj = ["Selecione",10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160, 200, 250, 320, 400, 500, 630]

        self.variable = tk.StringVar(self.JanelaCA)
        self.variable.set(disj[0])

        #someStyle = ttk.Style()
        #someStyle.configure('my.TMenubutton', font=("Arial","10","bold"), background='white')


        self.nome14 = ttk.Combobox(self.JanelaCA, textvariable=self.variable)
        self.nome14["width"] = 9
        #self.nome14["font"] = self.fontePadrao
        self.nome14.place(x=170, y=400)

        self.nome14['values'] = [disj[m] for m in range(0, 19)]

        # DIMENSIONAMENTO ELETRODUTO

        self.titulo = Label(self.JanelaCA, text="Dimensionamento eletroduto", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=420,y=175)

        self.nome12Label = Label(self.JanelaCA, text="N° cond. no eletroduto:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome12Label.place(x=400,y=205)

        self.nome12 = Entry(self.JanelaCA)
        self.nome12["width"] = 10
        self.nome12["font"] = self.fontePadrao
        self.nome12.place(x=540,y=205)

        self.nome13Label = Label(self.JanelaCA, text="Seção dos condutores:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome13Label.place(x=400,y=235)

        self.nome13 = Entry(self.JanelaCA)
        self.nome13["width"] = 10
        self.nome13["font"] = self.fontePadrao
        self.nome13.place(x=540,y=235)

        self.calcular3 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.calcular3["text"] = "Calcular"
        self.calcular3["font"] = ("Calibri", "8", "bold")
        self.calcular3["width"] = 12
        self.calcular3["command"] = self.Calcular3
        self.calcular3.place(x=450,y=265)

        self.mensagem3 = Label(self.JanelaCA, text="", font=self.fontePadrao, fg='deep sky blue', bg='gray18')
        self.mensagem3.place(x=400,y=295)

        elet = ["Selecione", 16, 20, 25, 32, 40, 50, 60, 75, 85]

        self.variable1 = tk.StringVar(self.JanelaCA)
        self.variable1.set(elet[0])

        self.nome15Label = Label(self.JanelaCA, text="Confirme o eletroduto: ", font=("Arial", "10", "bold", "italic"), fg='white', bg='gray18')
        self.nome15Label.place(x=400,y=325)

        self.nome15 = ttk.Combobox(self.JanelaCA, textvariable=self.variable1)
        self.nome15["width"] = 9
        #self.nome15["font"] = self.fontePadrao
        self.nome15.place(x=548,y=324)

        self.nome15['values'] = [elet[m] for m in range(0, 10)]

        # JANELAS COM AS TABELAS

        self.janela2 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.janela2["text"] = "Tabela Condutores PVC"
        self.janela2["font"] = ("Calibri", "8", "bold")
        self.janela2["width"] = 30
        self.janela2["command"] = self.tabela_condutores_PVC
        self.janela2.place(x=340, y=400)

        self.janela6 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.janela6["text"] = "Tabela Condutores EPR/XLPE"
        self.janela6["font"] = ("Calibri", "8", "bold")
        self.janela6["width"] = 30
        self.janela6["command"] = self.tabela_condutores_EPR_XLPE
        self.janela6.place(x=530, y=400)

        self.janela3 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.janela3["text"] = "Tabela Disjuntores"
        self.janela3["font"] = ("Calibri", "8", "bold")
        self.janela3["width"] = 30
        self.janela3["command"] = self.corrente_disjuntores
        self.janela3.place(x=340, y=423)

        self.janela4 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.janela4["text"] = "Fator de Correção de Temp."
        self.janela4["font"] = ("Calibri", "8", "bold")
        self.janela4["width"] = 30
        self.janela4["command"] = self.fator_correcao_temp
        self.janela4.place(x=530, y=423)

        self.janela5 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.janela5["text"] = "Fator de Correção de Agrup."
        self.janela5["font"] = ("Calibri", "8", "bold")
        self.janela5["width"] = 30
        self.janela5["command"] = self.fator_correcao_agrup
        self.janela5.place(x=340, y=446)

        self.janela7 = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.janela7["text"] = "Queda de tensão tabelada"
        self.janela7["font"] = ("Calibri", "8", "bold")
        self.janela7["width"] = 30
        self.janela7["command"] = self.queda_tensao_cond
        self.janela7.place(x=530, y=446)

        self.gerar = Button(self.JanelaCA, fg='darkorange2', bg='white')
        self.gerar["text"] = "Gerar lista"
        self.gerar["font"] = ("Calibri", "8", "bold")
        self.gerar["width"] = 30
        self.gerar["command"] = self.lista_materiais
        self.gerar.place(x=430, y=469)


    def Calcular(self):

        # DIMENSIONAMENTO CONFORME OS 5 CRITÉRIOS TÉCNICOS DA NBR 5410
        # CAPACIDADE DE CONDUÇÃO DE CORRENTE
        # SEÇÃO MINIMA DE 2,5 mm² PARA CIRCUITOS DE FORÇA E 1,5 mm² PARA CIRCUITOS DE ILUMINAÇÃO

        sec = [0.5, 0.75, 1, 1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240, 300, 400, 500, 630, 800, 1000]

        ip = float(self.nome2.get().replace(",", "."))
        fca = float(self.nome3.get().replace(",", "."))
        fct = float(self.nome4.get().replace(",", "."))
        met = self.nome5.get().upper()
        ncc = int(self.nome6.get())
        iso = self.nome7.get().upper()

        ic = ip/(fca*fct)

        # MÉTODO A1 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'A1' and ic > 0 and ic <= 7 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'A1' and ic > 7 and ic <= 9 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A1' and ic > 9 and ic <= 10 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A1' and ic > 10 and ic <= 13.5 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A1' and ic > 13.5 and ic <= 18 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A1' and ic > 18 and ic <= 24 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A1' and ic > 24 and ic <= 31 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A1' and ic > 31 and ic <= 42 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A1' and ic > 42 and ic <= 56 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A1' and ic > 56 and ic <= 73 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A1' and ic > 73 and ic <= 89 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A1' and ic > 89 and ic <= 108 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A1' and ic > 108 and ic <= 136 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A1' and ic > 136 and ic <= 164 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A1' and ic > 164 and ic <= 188 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A1' and ic > 188 and ic <= 216 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A1' and ic > 216 and ic <= 245 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A1' and ic > 245 and ic <= 286 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A1' and ic > 286 and ic <= 328 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A1' and ic > 328 and ic <= 390 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A1' and ic > 390 and ic <= 447 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A1' and ic > 447 and ic <= 514 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A1' and ic > 514 and ic <= 593 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A1' and ic > 593 and ic <= 679 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO A1 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'A1' and ic > 0 and ic <= 7 and ncc < 3 and iso == 'PVC':
            sec = 0.5
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec}"
        elif met == 'A1' and ic > 7 and ic <= 9 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A1' and ic > 9 and ic <= 11 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A1' and ic > 11 and ic <= 14.5 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A1' and ic > 14.5 and ic <= 19.5 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A1' and ic > 19.5 and ic <= 26 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A1' and ic > 26 and ic <= 34 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A1' and ic > 34 and ic <= 46 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A1' and ic > 46 and ic <= 61 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A1' and ic > 61 and ic <= 80 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A1' and ic > 80 and ic <= 99 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A1' and ic > 99 and ic <= 119 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A1' and ic > 119 and ic <= 151 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A1' and ic > 151 and ic <= 182 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A1' and ic > 182 and ic <= 210 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A1' and ic > 210 and ic <= 240 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A1' and ic > 240 and ic <= 273 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A1' and ic > 273 and ic <= 321 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A1' and ic > 321 and ic <= 367 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A1' and ic > 367 and ic <= 438 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A1' and ic > 438 and ic <= 502 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A1' and ic > 502 and ic <= 578 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A1' and ic > 578 and ic <= 669 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A1' and ic > 669 and ic <= 767 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"


        # MÉTODO A2 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'A2' and ic > 0 and ic <= 7 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'A2' and ic > 7 and ic <= 9 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A2' and ic > 9 and ic <= 10 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A2' and ic > 10 and ic <= 13 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A2' and ic > 13 and ic <= 17.5 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A2' and ic > 17.5 and ic <= 23 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A2' and ic > 23 and ic <= 29 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A2' and ic > 29 and ic <= 39 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A2' and ic > 39 and ic <= 52 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A2' and ic > 52 and ic <= 68 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A2' and ic > 68 and ic <= 83 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A2' and ic > 83 and ic <= 99 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A2' and ic > 99 and ic <= 125 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A2' and ic > 125 and ic <= 150 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A2' and ic > 150 and ic <= 172 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A2' and ic > 172 and ic <= 196 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A2' and ic > 196 and ic <= 223 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A2' and ic > 223 and ic <= 261 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A2' and ic > 261 and ic <= 298 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A2' and ic > 298 and ic <= 355 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A2' and ic > 355 and ic <= 406 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A2' and ic > 406 and ic <= 467 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A2' and ic > 467 and ic <= 540 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A2' and ic > 540 and ic <= 618 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO A2 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'A2' and ic > 0 and ic <= 7 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'A2' and ic > 7 and ic <= 9 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A2' and ic > 9 and ic <= 11 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A2' and ic > 11 and ic <= 14 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A2' and ic > 14 and ic <= 18.5 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A2' and ic > 18.5 and ic <= 25 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A2' and ic > 25 and ic <= 32 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A2' and ic > 32 and ic <= 43 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A2' and ic > 43 and ic <= 57 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A2' and ic > 57 and ic <= 75 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A2' and ic > 75 and ic <= 92 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A2' and ic > 92 and ic <= 110 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A2' and ic > 110 and ic <= 139 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A2' and ic > 139 and ic <= 167 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A2' and ic > 167 and ic <= 192 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A2' and ic > 192 and ic <= 219 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A2' and ic > 219 and ic <= 248 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A2' and ic > 248 and ic <= 291 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A2' and ic > 291 and ic <= 334 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A2' and ic > 334 and ic <= 398 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A2' and ic > 398 and ic <= 456 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A2' and ic > 456 and ic <= 526 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A2' and ic > 526 and ic <= 609 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A2' and ic > 609 and ic <= 698 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B1 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'B1' and ic > 0 and ic <= 8 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B1' and ic > 8 and ic <= 10 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B1' and ic > 10 and ic <= 12 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B1' and ic > 12 and ic <= 15.5 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B1' and ic > 15.5 and ic <= 21 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B1' and ic > 21 and ic <= 28 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B1' and ic > 28 and ic <= 36 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B1' and ic > 36 and ic <= 50 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B1' and ic > 50 and ic <= 68 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B1' and ic > 68 and ic <= 89 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B1' and ic > 89 and ic <= 110 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B1' and ic > 110 and ic <= 134 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B1' and ic > 134 and ic <= 171 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B1' and ic > 171 and ic <= 207 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B1' and ic > 207 and ic <= 239 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B1' and ic > 239 and ic <= 275 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B1' and ic > 275 and ic <= 314 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B1' and ic > 314 and ic <= 370 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B1' and ic > 370 and ic <= 426 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B1' and ic > 426 and ic <= 510 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B1' and ic > 510 and ic <= 587 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B1' and ic > 587 and ic <= 678 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B1' and ic > 678 and ic <= 788 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B1' and ic > 788 and ic <= 906 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B1 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'B1' and ic > 0 and ic <= 9 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B1' and ic > 9 and ic <= 11 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B1' and ic > 11 and ic <= 14 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B1' and ic > 14 and ic <= 17.5 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B1' and ic > 17.5 and ic <= 24 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B1' and ic > 24 and ic <= 32 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B1' and ic > 32 and ic <= 41 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B1' and ic > 41 and ic <= 57 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B1' and ic > 57 and ic <= 76 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B1' and ic > 76 and ic <= 101 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B1' and ic > 101 and ic <= 125 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B1' and ic > 125 and ic <= 151 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B1' and ic > 151 and ic <= 192 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B1' and ic > 192 and ic <= 232 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B1' and ic > 232 and ic <= 269 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B1' and ic > 269 and ic <= 309 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B1' and ic > 309 and ic <= 353 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B1' and ic > 353 and ic <= 415 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B1' and ic > 415 and ic <= 477 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B1' and ic > 477 and ic <= 571 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B1' and ic > 571 and ic <= 656 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B1' and ic > 656 and ic <= 758 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B1' and ic > 758 and ic <= 881 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B1' and ic > 881 and ic <= 1012 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B2 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'B2' and ic > 0 and ic <= 8 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B2' and ic > 8 and ic <= 10 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B2' and ic > 10 and ic <= 12 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B2' and ic > 12 and ic <= 15 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B2' and ic > 15 and ic <= 20 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B2' and ic > 20 and ic <= 27 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B2' and ic > 27 and ic <= 34 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B2' and ic > 34 and ic <= 46 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B2' and ic > 46 and ic <= 62 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B2' and ic > 62 and ic <= 80 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B2' and ic > 80 and ic <= 99 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B2' and ic > 99 and ic <= 118 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B2' and ic > 118 and ic <= 149 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B2' and ic > 149 and ic <= 179 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B2' and ic > 179 and ic <= 206 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B2' and ic > 206 and ic <= 236 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B2' and ic > 236 and ic <= 268 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B2' and ic > 268 and ic <= 313 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B2' and ic > 313 and ic <= 358 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B2' and ic > 358 and ic <= 425 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B2' and ic > 425 and ic <= 486 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B2' and ic > 486 and ic <= 559 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B2' and ic > 559 and ic <= 645 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B2' and ic > 645 and ic <= 738 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B2 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'B2' and ic > 0 and ic <= 9 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B2' and ic > 9 and ic <= 11 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B2' and ic > 11 and ic <= 13 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B2' and ic > 13 and ic <= 16.5 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B2' and ic > 16.5 and ic <= 23 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B2' and ic > 23 and ic <= 30 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B2' and ic > 30 and ic <= 38 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B2' and ic > 38 and ic <= 52 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B2' and ic > 52 and ic <= 69 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B2' and ic > 69 and ic <= 90 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B2' and ic > 90 and ic <= 111 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B2' and ic > 111 and ic <= 133 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B2' and ic > 133 and ic <= 168 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B2' and ic > 168 and ic <= 201 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B2' and ic > 201 and ic <= 232 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B2' and ic > 232 and ic <= 265 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B2' and ic > 265 and ic <= 300 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B2' and ic > 300 and ic <= 351 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B2' and ic > 351 and ic <= 401 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B2' and ic > 401 and ic <= 477 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B2' and ic > 477 and ic <= 545 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B2' and ic > 545 and ic <= 626 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B2' and ic > 626 and ic <= 723 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B2' and ic > 723 and ic <= 827 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO C PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'C' and ic > 0 and ic <= 9 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'C' and ic > 9 and ic <= 11 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'C' and ic > 11 and ic <= 14 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'C' and ic > 14 and ic <= 17.5 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'C' and ic > 17.5 and ic <= 24 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'C' and ic > 24 and ic <= 32 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'C' and ic > 32 and ic <= 41 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'C' and ic > 41 and ic <= 57 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'C' and ic > 57 and ic <= 76 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'C' and ic > 76 and ic <= 96 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'C' and ic > 96 and ic <= 119 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'C' and ic > 119 and ic <= 144 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'C' and ic > 144 and ic <= 184 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'C' and ic > 184 and ic <= 223 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'C' and ic > 223 and ic <= 259 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'C' and ic > 259 and ic <= 299 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'C' and ic > 299 and ic <= 341 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'C' and ic > 341 and ic <= 403 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'C' and ic > 403 and ic <= 464 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'C' and ic > 464 and ic <= 557 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'C' and ic > 557 and ic <= 642 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'C' and ic > 642 and ic <= 743 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'C' and ic > 743 and ic <= 865 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'C' and ic > 865 and ic <= 996 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO C PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'C' and ic > 0 and ic <= 10 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'C' and ic > 10 and ic <= 13 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'C' and ic > 13 and ic <= 15 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'C' and ic > 15 and ic <= 19.5 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'C' and ic > 19.5 and ic <= 27 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'C' and ic > 27 and ic <= 36 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'C' and ic > 36 and ic <= 46 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'C' and ic > 46 and ic <= 63 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'C' and ic > 63 and ic <= 85 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'C' and ic > 85 and ic <= 112 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'C' and ic > 112 and ic <= 138 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'C' and ic > 138 and ic <= 168 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'C' and ic > 168 and ic <= 213 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'C' and ic > 213 and ic <= 258 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'C' and ic > 258 and ic <= 299 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'C' and ic > 299 and ic <= 344 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'C' and ic > 344 and ic <= 392 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'C' and ic > 392 and ic <= 461 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'C' and ic > 461 and ic <= 530 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'C' and ic > 530 and ic <= 634 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'C' and ic > 634 and ic <= 729 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'C' and ic > 729 and ic <= 843 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'C' and ic > 843 and ic <= 978 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'C' and ic > 978 and ic <= 1125 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO D PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'D' and ic > 0 and ic <= 10 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'D' and ic > 10 and ic <= 12 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'D' and ic > 12 and ic <= 15 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'D' and ic > 15 and ic <= 18 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'D' and ic > 18 and ic <= 24 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'D' and ic > 24 and ic <= 31 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'D' and ic > 31 and ic <= 39 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'D' and ic > 39 and ic <= 52 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'D' and ic > 52 and ic <= 67 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'D' and ic > 67 and ic <= 86 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'D' and ic > 86 and ic <= 103 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'D' and ic > 103 and ic <= 122 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'D' and ic > 122 and ic <= 151 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'D' and ic > 151 and ic <= 179 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'D' and ic > 179 and ic <= 203 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'D' and ic > 203 and ic <= 230 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'D' and ic > 230 and ic <= 258 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'D' and ic > 258 and ic <= 297 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'D' and ic > 297 and ic <= 336 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'D' and ic > 336 and ic <= 394 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'D' and ic > 394 and ic <= 445 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'D' and ic > 445 and ic <= 506 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'D' and ic > 506 and ic <= 577 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'D' and ic > 577 and ic <= 652 and ncc > 2 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO D PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO PVC
        if met == 'D' and ic > 0 and ic <= 12 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'D' and ic > 12 and ic <= 15 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'D' and ic > 15 and ic <= 18 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'D' and ic > 18 and ic <= 22 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'D' and ic > 22 and ic <= 29 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'D' and ic > 29 and ic <= 38 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'D' and ic > 38 and ic <= 47 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'D' and ic > 47 and ic <= 63 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'D' and ic > 63 and ic <= 81 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'D' and ic > 81 and ic <= 104 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'D' and ic > 104 and ic <= 125 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'D' and ic > 125 and ic <= 148 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'D' and ic > 148 and ic <= 183 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'D' and ic > 183 and ic <= 216 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'D' and ic > 216 and ic <= 246 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'D' and ic > 246 and ic <= 278 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'D' and ic > 278 and ic <= 312 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'D' and ic > 312 and ic <= 361 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'D' and ic > 361 and ic <= 408 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'D' and ic > 408 and ic <= 478 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'D' and ic > 478 and ic <= 540 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'D' and ic > 540 and ic <= 614 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'D' and ic > 614 and ic <= 700 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'D' and ic > 700 and ic <= 792 and ncc < 3 and iso == 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"





########################################################################################################################




        # MÉTODO A1 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'A1' and ic > 0 and ic <= 9 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'A1' and ic > 9 and ic <= 11 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A1' and ic > 11 and ic <= 13 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A1' and ic > 13 and ic <= 17 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A1' and ic > 17 and ic <= 23 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A1' and ic > 23 and ic <= 31 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A1' and ic > 31 and ic <= 40 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A1' and ic > 40 and ic <= 54 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A1' and ic > 54 and ic <= 73 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A1' and ic > 73 and ic <= 95 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A1' and ic > 95 and ic <= 117 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A1' and ic > 117 and ic <= 141 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A1' and ic > 141 and ic <= 179 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A1' and ic > 179 and ic <= 216 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A1' and ic > 216 and ic <= 249 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A1' and ic > 249 and ic <= 285 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A1' and ic > 285 and ic <= 324 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A1' and ic > 324 and ic <= 380 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A1' and ic > 380 and ic <= 435 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A1' and ic > 435 and ic <= 519 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A1' and ic > 519 and ic <= 595 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A1' and ic > 595 and ic <= 685 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A1' and ic > 685 and ic <= 792 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A1' and ic > 792 and ic <= 908 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO A1 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'A1' and ic > 0 and ic <= 10 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'A1' and ic > 10 and ic <= 12 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A1' and ic > 12 and ic <= 15 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A1' and ic > 15 and ic <= 19 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A1' and ic > 19 and ic <= 26 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A1' and ic > 26 and ic <= 35 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A1' and ic > 35 and ic <= 45 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A1' and ic > 45 and ic <= 61 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A1' and ic > 61 and ic <= 81 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A1' and ic > 81 and ic <= 106 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A1' and ic > 106 and ic <= 131 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A1' and ic > 131 and ic <= 158 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A1' and ic > 158 and ic <= 200 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A1' and ic > 200 and ic <= 241 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A1' and ic > 241 and ic <= 278 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A1' and ic > 278 and ic <= 318 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A1' and ic > 318 and ic <= 362 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A1' and ic > 362 and ic <= 424 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A1' and ic > 424 and ic <= 486 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A1' and ic > 486 and ic <= 579 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A1' and ic > 579 and ic <= 664 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A1' and ic > 664 and ic <= 765 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A1' and ic > 765 and ic <= 885 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A1' and ic > 885 and ic <= 1014 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO A2 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'A2' and ic > 0 and ic <= 9 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'A2' and ic > 9 and ic <= 11 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A2' and ic > 11 and ic <= 13 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A2' and ic > 13 and ic <= 16.5 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A2' and ic > 16.5 and ic <= 22 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A2' and ic > 22 and ic <= 30 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A2' and ic > 30 and ic <= 38 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A2' and ic > 38 and ic <= 51 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A2' and ic > 51 and ic <= 68 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A2' and ic > 68 and ic <= 89 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A2' and ic > 89 and ic <= 109 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A2' and ic > 109 and ic <= 130 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A2' and ic > 130 and ic <= 164 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A2' and ic > 164 and ic <= 197 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A2' and ic > 197 and ic <= 227 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A2' and ic > 227 and ic <= 259 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A2' and ic > 259 and ic <= 295 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A2' and ic > 295 and ic <= 346 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A2' and ic > 346 and ic <= 396 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A2' and ic > 396 and ic <= 472 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A2' and ic > 472 and ic <= 541 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A2' and ic > 541 and ic <= 623 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A2' and ic > 623 and ic <= 721 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A2' and ic > 721 and ic <= 826 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO A2 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'A2' and ic > 0 and ic <= 10 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'A2' and ic > 10 and ic <= 12 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'A2' and ic > 12 and ic <= 14 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'A2' and ic > 14 and ic <= 18.5 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'A2' and ic > 18.5 and ic <= 25 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'A2' and ic > 25 and ic <= 33 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'A2' and ic > 33 and ic <= 42 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'A2' and ic > 42 and ic <= 57 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'A2' and ic > 57 and ic <= 76 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'A2' and ic > 76 and ic <= 99 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'A2' and ic > 99 and ic <= 121 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'A2' and ic > 121 and ic <= 145 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'A2' and ic > 145 and ic <= 183 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'A2' and ic > 183 and ic <= 220 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'A2' and ic > 220 and ic <= 253 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'A2' and ic > 253 and ic <= 290 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'A2' and ic > 290 and ic <= 329 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'A2' and ic > 329 and ic <= 386 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'A2' and ic > 386 and ic <= 442 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'A2' and ic > 442 and ic <= 527 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'A2' and ic > 527 and ic <= 604 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'A2' and ic > 604 and ic <= 696 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'A2' and ic > 696 and ic <= 805 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'A2' and ic > 805 and ic <= 923 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B1 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'B1' and ic > 0 and ic <= 10 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B1' and ic > 10 and ic <= 13 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B1' and ic > 13 and ic <= 16 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B1' and ic > 16 and ic <= 20 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B1' and ic > 20 and ic <= 28 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B1' and ic > 28 and ic <= 37 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B1' and ic > 37 and ic <= 48 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B1' and ic > 48 and ic <= 66 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B1' and ic > 66 and ic <= 88 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B1' and ic > 88 and ic <= 117 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B1' and ic > 117 and ic <= 144 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B1' and ic > 144 and ic <= 175 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B1' and ic > 175 and ic <= 222 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B1' and ic > 222 and ic <= 269 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B1' and ic > 269 and ic <= 312 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B1' and ic > 312 and ic <= 358 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B1' and ic > 358 and ic <= 408 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B1' and ic > 408 and ic <= 481 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B1' and ic > 481 and ic <= 553 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B1' and ic > 553 and ic <= 661 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B1' and ic > 661 and ic <= 760 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B1' and ic > 760 and ic <= 879 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B1' and ic > 879 and ic <= 1020 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B1' and ic > 1020 and ic <= 1173 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B1 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'B1' and ic > 0 and ic <= 12 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B1' and ic > 12 and ic <= 15 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B1' and ic > 15 and ic <= 18 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B1' and ic > 18 and ic <= 23 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B1' and ic > 23 and ic <= 31 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B1' and ic > 31 and ic <= 42 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B1' and ic > 42 and ic <= 54 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B1' and ic > 54 and ic <= 75 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B1' and ic > 75 and ic <= 100 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B1' and ic > 100 and ic <= 133 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B1' and ic > 133 and ic <= 164 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B1' and ic > 164 and ic <= 198 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B1' and ic > 198 and ic <= 253 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B1' and ic > 253 and ic <= 306 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B1' and ic > 306 and ic <= 354 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B1' and ic > 354 and ic <= 407 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B1' and ic > 407 and ic <= 464 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B1' and ic > 464 and ic <= 546 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B1' and ic > 546 and ic <= 628 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B1' and ic > 628 and ic <= 751 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B1' and ic > 751 and ic <= 864 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B1' and ic > 864 and ic <= 998 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B1' and ic > 998 and ic <= 1158 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B1' and ic > 1158 and ic <= 1332 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B2 PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'B2' and ic > 0 and ic <= 10 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B2' and ic > 10 and ic <= 13 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B2' and ic > 13 and ic <= 15 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B2' and ic > 15 and ic <= 19.5 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B2' and ic > 19.5 and ic <= 26 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B2' and ic > 26 and ic <= 35 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B2' and ic > 35 and ic <= 44 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B2' and ic > 44 and ic <= 60 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B2' and ic > 60 and ic <= 80 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B2' and ic > 80 and ic <= 105 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B2' and ic > 105 and ic <= 128 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B2' and ic > 128 and ic <= 154 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B2' and ic > 154 and ic <= 194 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B2' and ic > 194 and ic <= 233 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B2' and ic > 233 and ic <= 268 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B2' and ic > 268 and ic <= 307 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B2' and ic > 307 and ic <= 348 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B2' and ic > 348 and ic <= 407 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B2' and ic > 407 and ic <= 465 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B2' and ic > 465 and ic <= 552 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B2' and ic > 552 and ic <= 631 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B2' and ic > 631 and ic <= 725 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B2' and ic > 725 and ic <= 837 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B2' and ic > 837 and ic <= 957 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO B2 PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'B2' and ic > 0 and ic <= 11 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'B2' and ic > 11 and ic <= 15 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'B2' and ic > 15 and ic <= 17 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'B2' and ic > 17 and ic <= 22 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'B2' and ic > 22 and ic <= 30 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'B2' and ic > 30 and ic <= 40 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'B2' and ic > 40 and ic <= 51 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'B2' and ic > 51 and ic <= 69 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'B2' and ic > 69 and ic <= 91 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'B2' and ic > 91 and ic <= 119 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'B2' and ic > 119 and ic <= 146 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'B2' and ic > 146 and ic <= 175 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'B2' and ic > 175 and ic <= 221 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'B2' and ic > 221 and ic <= 265 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'B2' and ic > 265 and ic <= 305 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'B2' and ic > 305 and ic <= 349 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'B2' and ic > 349 and ic <= 395 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'B2' and ic > 395 and ic <= 462 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'B2' and ic > 462 and ic <= 529 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'B2' and ic > 529 and ic <= 628 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'B2' and ic > 628 and ic <= 718 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'B2' and ic > 718 and ic <= 825 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'B2' and ic > 825 and ic <= 952 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'B2' and ic > 952 and ic <= 1088 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO C PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'C' and ic > 0 and ic <= 11 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'C' and ic > 11 and ic <= 14 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'C' and ic > 14 and ic <= 17 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'C' and ic > 17 and ic <= 22 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'C' and ic > 22 and ic <= 30 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'C' and ic > 30 and ic <= 40 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'C' and ic > 40 and ic <= 52 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'C' and ic > 52 and ic <= 71 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'C' and ic > 71 and ic <= 96 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'C' and ic > 96 and ic <= 119 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'C' and ic > 119 and ic <= 147 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'C' and ic > 147 and ic <= 179 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'C' and ic > 179 and ic <= 229 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'C' and ic > 229 and ic <= 278 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'C' and ic > 278 and ic <= 322 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'C' and ic > 322 and ic <= 371 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'C' and ic > 371 and ic <= 424 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'C' and ic > 424 and ic <= 500 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'C' and ic > 500 and ic <= 576 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'C' and ic > 576 and ic <= 692 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'C' and ic > 692 and ic <= 797 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'C' and ic > 797 and ic <= 923 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'C' and ic > 923 and ic <= 1074 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'C' and ic > 1074 and ic <= 1237 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO C PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'C' and ic > 0 and ic <= 12 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'C' and ic > 12 and ic <= 16 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'C' and ic > 16 and ic <= 19 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'C' and ic > 19 and ic <= 24 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'C' and ic > 24 and ic <= 33 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'C' and ic > 33 and ic <= 45 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'C' and ic > 45 and ic <= 58 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'C' and ic > 58 and ic <= 80 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'C' and ic > 80 and ic <= 107 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'C' and ic > 107 and ic <= 138 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'C' and ic > 138 and ic <= 171 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'C' and ic > 171 and ic <= 209 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'C' and ic > 209 and ic <= 269 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'C' and ic > 269 and ic <= 328 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'C' and ic > 328 and ic <= 382 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'C' and ic > 382 and ic <= 441 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'C' and ic > 441 and ic <= 506 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'C' and ic > 506 and ic <= 599 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'C' and ic > 599 and ic <= 693 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'C' and ic > 693 and ic <= 835 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'C' and ic > 835 and ic <= 966 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'C' and ic > 966 and ic <= 1122 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'C' and ic > 1122 and ic <= 1311 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'C' and ic > 1311 and ic <= 1515 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO D PARA 3 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'D' and ic > 0 and ic <= 12 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'D' and ic > 12 and ic <= 15 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'D' and ic > 15 and ic <= 17 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'D' and ic > 17 and ic <= 22 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'D' and ic > 22 and ic <= 29 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'D' and ic > 29 and ic <= 37 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'D' and ic > 37 and ic <= 46 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'D' and ic > 46 and ic <= 61 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'D' and ic > 61 and ic <= 79 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'D' and ic > 79 and ic <= 101 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'D' and ic > 101 and ic <= 122 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'D' and ic > 122 and ic <= 144 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'D' and ic > 144 and ic <= 178 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'D' and ic > 178 and ic <= 211 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'D' and ic > 211 and ic <= 240 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'D' and ic > 240 and ic <= 271 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'D' and ic > 271 and ic <= 304 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'D' and ic > 304 and ic <= 351 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'D' and ic > 351 and ic <= 396 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'D' and ic > 396 and ic <= 464 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'D' and ic > 464 and ic <= 525 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'D' and ic > 525 and ic <= 596 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'D' and ic > 596 and ic <= 679 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'D' and ic > 679 and ic <= 767 and ncc > 2 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"

        # MÉTODO D PARA 2 CONDUTORES CARREGADOS ISOLAÇÃO EPR OU XLPE
        if met == 'D' and ic > 0 and ic <= 14 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[0]}"
        elif met == 'D' and ic > 14 and ic <= 18 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[1]}"
        elif met == 'D' and ic > 18 and ic <= 21 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[2]}"
        elif met == 'D' and ic > 21 and ic <= 26 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[3]}"
        elif met == 'D' and ic > 26 and ic <= 34 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[4]}"
        elif met == 'D' and ic > 34 and ic <= 44 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[5]}"
        elif met == 'D' and ic > 44 and ic <= 56 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[6]}"
        elif met == 'D' and ic > 56 and ic <= 73 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[7]}"
        elif met == 'D' and ic > 73 and ic <= 95 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[8]}"
        elif met == 'D' and ic > 95 and ic <= 121 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[9]}"
        elif met == 'D' and ic > 121 and ic <= 146 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[10]}"
        elif met == 'D' and ic > 146 and ic <= 173 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[11]}"
        elif met == 'D' and ic > 173 and ic <= 213 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[12]}"
        elif met == 'D' and ic > 213 and ic <= 252 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[13]}"
        elif met == 'D' and ic > 252 and ic <= 287 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[14]}"
        elif met == 'D' and ic > 287 and ic <= 324 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[15]}"
        elif met == 'D' and ic > 324 and ic <= 363 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[16]}"
        elif met == 'D' and ic > 363 and ic <= 419 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[17]}"
        elif met == 'D' and ic > 419 and ic <= 474 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[18]}"
        elif met == 'D' and ic > 474 and ic <= 555 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[19]}"
        elif met == 'D' and ic > 555 and ic <= 627 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[20]}"
        elif met == 'D' and ic > 627 and ic <= 711 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[21]}"
        elif met == 'D' and ic > 711 and ic <= 811 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[22]}"
        elif met == 'D' and ic > 811 and ic <= 916 and ncc < 3 and iso != 'PVC':
            self.mensagem["font"] = ("Arial", "10", "bold")
            self.mensagem["text"] = f"A seção do condutor deverá ser em mm²: {sec[23]}"


    def Calcular1(self):

        # DIMENSIONAMENTO CONFORME OS 5 CRITÉRIOS TÉCNICOS DA NBR 5410
        # CRITÉRIO DE QUEDA DE TENSÃO
        # QUEDA DE TENSÃO MÁXIMA EM CIRCUITOS TERMINAIS É DE 4%

        v = float(self.nome8.get().replace(",", "."))
        ip = float(self.nome2.get().replace(",", "."))
        qdc = float(self.nome9.get().replace(",", "."))
        l = float(self.nome10.get().replace(",", "."))

        du = qdc*ip*l
        du1 = (du / v) * 100
        if du1 < 4:
            self.mensagem1["foreground"] = "deep sky blue"
            self.mensagem1["font"] = ("Arial", "10", "bold")
            self.mensagem1["text"] = "Critério de queda de tensão atendido!"
        else:
            self.mensagem1["foreground"] = "red"
            self.mensagem1["font"] = ("Arial", "10", "bold")
            self.mensagem1["text"] = """Critério de queda de tensão NÃO atendido!"""
#rever as condições do circuito!"""


    def Calcular2(self):

        # DIMENSIONAMENTO CONFORME OS 5 CRITÉRIOS TÉCNICOS DA NBR 5410
        # CRITÉRIO DE SOBRECARGA


        dj = [10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160, 200, 250, 320, 400, 500, 630]

        ip = float(self.nome2.get().replace(",", "."))
        ic1 = float(self.nome11.get().replace(",", "."))
        fca = float(self.nome3.get().replace(",", "."))
        fct = float(self.nome4.get().replace(",", "."))

        iz = ic1 * fca * fct
        it = 0
        for dado in dj:
            if ip < dado <= iz:
                it = dado
                self.mensagem2["foreground"] = "deep sky blue"
                self.mensagem2["font"] = ("Arial", "10", "bold")
                self.mensagem2["text"] = f"O disjuntor deverá possuir corrente nominal de {it} A"
            elif it == 0:
                self.mensagem2["foreground"] = "red"
                self.mensagem2["font"] = ("Arial", "10", "bold")
                self.mensagem2["text"] = """Não existe valor comercial de disjuntor para esta faixa de corrente,\naumente a seção do condutor e recalcule!"""


    def Calcular3(self):

        # DIMENSIONAMENTO CONFORME OS 5 CRITÉRIOS TÉCNICOS DA NBR 5410
        # DIMENSIONAMENTO DE ELETRODUTO

        ne = int(self.nome12.get())
        sec = float(self.nome13.get().replace(",", "."))

        if ne > 1 and ne < 8 and sec == 1.5:
            te = 16
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 5 and sec == 2.5:
            te = 16
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 4 and sec == 4:
            te = 16
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 6:
            te = 16
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 7 and ne < 11 and sec == 1.5:
            te = 20
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 9 and sec == 2.5:
            te = 20
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 7 and sec == 4:
            te = 20
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 5 and sec == 6:
            te = 20
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 4 and sec == 10:
            te = 20
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 16:
            te = 20
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 8 and ne < 11 and sec == 2.5:
            te = 25
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 6 and ne < 11 and sec == 4:
            te = 25
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 9 and sec == 6:
            te = 25
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 6 and sec == 10:
            te = 25
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 5 and sec == 16:
            te = 25
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 25:
            te = 25
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 35:
            te = 25
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 8 and ne < 11 and sec == 6:
            te = 32
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 5 and ne < 9 and sec == 10:
            te = 32
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 7 and sec == 16:
            te = 32
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 5 and sec == 25:
            te = 32
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 4 and sec == 35:
            te = 32
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 50:
            te = 32
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 8 and ne < 11 and sec == 10:
            te = 40
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 6 and ne < 11 and sec == 16:
            te = 40
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 8 and sec == 25:
            te = 40
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 6 and sec == 35:
            te = 40
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 5 and sec == 50:
            te = 40
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 4 and sec == 70:
            te = 40
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 95:
            te = 40
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 7 and ne < 11 and sec == 25:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 5 and ne < 10 and sec == 35:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 7 and sec == 50:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 6 and sec == 70:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 4 and sec == 95:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 4 and sec == 120:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 150:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 185:
            te = 50
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 9 and ne < 11 and sec == 35:
            te = 60
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 6 and ne < 10 and sec == 50:
            te = 60
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 8 and sec == 70:
            te = 60
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 6 and sec == 95:
            te = 60
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 5 and sec == 120:
            te = 60
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 4 and sec == 150:
            te = 60
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 1 and ne < 3 and sec == 240:
            te = 60
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 9 and ne < 11 and sec == 50:
            te = 75
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 7 and ne < 11 and sec == 70:
            te = 75
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 5 and ne < 9 and sec == 95:
            te = 75
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 8 and sec == 120:
            te = 75
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 6 and sec == 150:
            te = 75
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 5 and sec == 185:
            te = 75
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 2 and ne < 4 and sec == 240:
            te = 75
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 8 and ne < 11 and sec == 95:
            te = 85
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 7 and ne < 10 and sec == 120:
            te = 85
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 5 and ne < 8 and sec == 150:
            te = 85
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 4 and ne < 7 and sec == 185:
            te = 85
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"
        elif ne > 3 and ne < 5 and sec == 240:
            te = 85
            self.mensagem3["font"] = ("Arial", "10", "bold")
            self.mensagem3["text"] = f"O tamanho nominal do eletroduto deve ser em mm: {te}"

    def tabela_condutores_PVC(self):

        self.janela2 = Toplevel()
        self.janela2.title("Tabela de condutores PVC")
        self.janela2.iconbitmap('baixados.ico')
        self.imag = PhotoImage(file="tabela_condutores_PVC.png")
        self.figura = Label(self.janela2, image=self.imag)
        self.figura.pack()

    def tabela_condutores_EPR_XLPE(self):

        self.janela6 = Toplevel()
        #self.janela6.geometry('1000x1000')
        self.janela6.title("Tabela de condutores EPR/XLPE")
        self.janela6.iconbitmap('baixados.ico')
        self.imag4 = PhotoImage(file="tabela_condutores_EPR_XLPE.png")
        self.figura4 = Label(self.janela6, image=self.imag4)
        self.figura4.pack()

    def corrente_disjuntores(self):

        self.janela3 = Toplevel()
        self.janela3.title("Corrente Disjuntores comerciais")
        self.janela3.iconbitmap('baixados.ico')
        self.imag1 = PhotoImage(file="disjuntores_comerciais1.png")
        self.figura1 = Label(self.janela3, image=self.imag1)
        self.figura1.pack()

    def fator_correcao_temp(self):

        self.janela4 = Toplevel()
        self.janela4.title("Fatores de correção de temperatura")
        self.janela4.iconbitmap('baixados.ico')
        self.imag2 = PhotoImage(file="fator_correcao_temperatura.png")
        self.figura2 = Label(self.janela4, image=self.imag2)
        self.figura2.pack()

    def fator_correcao_agrup(self):

        self.janela5 = Toplevel()
        self.janela5.title("Fatores de correção de agrupamento")
        self.janela5.iconbitmap('baixados.ico')
        self.imag3 = PhotoImage(file="fator_correcao_agrupamento.png")
        self.figura3 = Label(self.janela5, image=self.imag3)
        self.figura3.pack()

    def queda_tensao_cond(self):

        self.janela7 = Toplevel()
        self.janela7.title("Queda de tensão tabelada")
        self.janela7.iconbitmap('baixados.ico')
        self.imag5 = PhotoImage(file="queda_tensao_cond.png")
        self.figura5 = Label(self.janela7, image=self.imag5)
        self.figura5.pack()

    def lista_materiais(self):

        # GERA LISTA DE MATERIAIS

        ncc = int(self.nome6.get())
        l = float(self.nome10.get().replace(",", "."))
        sec = float(self.nome13.get().replace(",", "."))
        dsj = float(self.variable.get())
        te = int(self.variable1.get())

        self.gerar = xlsxwriter.Workbook('Lista de Materiais CA.xlsx')
        self.worksheet = self.gerar.add_worksheet('Lista')

        dados = [
            ["Produto", "Característica", "Quantidade"],
            ["Condutor Fase", f"{sec} mm²", f"{((l * 1000) * ncc)+(0.1*((l * 1000) * ncc))} m"],
            ["Condutor Neutro", f"{sec} mm²", f"{(l * 1000)+(0.1*(l * 1000))} m"],
            ["Condutor Terra", f"{sec} mm²", f"{(l * 1000)+(0.1*(l * 1000))} m"],
            ["Disjuntor", f"{dsj} A, {ncc}P", 1],
            ["Eletroduto", f"{te} mm", f"{(l * 1000)+(0.1*(l * 1000))} m"],
        ]
        row = 0
        col = 0
        for prod, secao, qtde in (dados):
            self.worksheet.write(row, col, prod)
            self.worksheet.write(row, col + 1, secao)
            self.worksheet.write(row, col + 2, qtde)
            row += 1

        self.gerar.close()








    # DIMENSIONAMENTO FV CONFORME A NBR 16612

    def JanelaCC(self):

        self.JanelaCC = Toplevel()
        self.JanelaCC.geometry("800x580")
        self.JanelaCC.minsize(800, 580)
        self.JanelaCC.maxsize(800, 580)
        self.JanelaCC.configure(bg='gray18')
        self.JanelaCC.title("Dimensionamento CC")
        self.JanelaCC.iconbitmap('baixados.ico')

        self.fontePadrao = ("Arial", "10")


        # CAPACIDADE DE CONDUÇÃO DE CORRENTE
        # ADOTAR SEÇÃO MINIMA DE 4mm²


        self.titulo = Label(self.JanelaCC, text="Capacidade de condução de corrente", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=80, y=5)



        self.nome17Label = Label(self.JanelaCC, text="Corrente de operação do módulo (NMOT):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome17Label.place(x=60, y=35)

        self.nome17 = Entry(self.JanelaCC)
        self.nome17["width"] = 10
        self.nome17["font"] = self.fontePadrao
        self.nome17.place(x=305, y=35)




        self.nome20Label = Label(self.JanelaCC, text="Fator de agrupamento:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome20Label.place(x=60, y=65)

        self.nome20 = Entry(self.JanelaCC)
        self.nome20["width"] = 10
        self.nome20["font"] = self.fontePadrao
        self.nome20.place(x=305, y=65)




        self.nome21Label = Label(self.JanelaCC, text="Método:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome21Label.place(x=60, y=95)

        metodoCC = ['Selecione', 'C1', 'C2', 'C3', 'C4']

        self.variable4 = tk.StringVar(self.JanelaCC)
        self.variable4.set(metodoCC[0])

        self.nome21 = ttk.Combobox(self.JanelaCC, textvariable=self.variable4)
        self.nome21["width"] = 9
        self.nome21.place(x=305, y=95)

        self.nome21['values'] = [metodoCC[m] for m in range(0, 5)]


        self.nome23Label = Label(self.JanelaCC, text="Se C1, modo:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome23Label.place(x=60, y=125)

        modoCC = ['Selecione', '1', '2', '3', '4']

        self.variable5 = tk.StringVar(self.JanelaCC)
        self.variable5.set(modoCC[0])

        self.nome23 = ttk.Combobox(self.JanelaCC, textvariable=self.variable5)
        self.nome23["width"] = 9
        self.nome23.place(x=305, y=125)

        self.nome23['values'] = [modoCC[m] for m in range(0, 5)]



        self.nome22Label = Label(self.JanelaCC, text="Se C1, exposto ao sol?", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome22Label.place(x=60, y=155)

        es = ['Selecione', 'S', 'N']

        self.variable6 = tk.StringVar(self.JanelaCC)
        self.variable6.set(es[0])

        self.nome22 = ttk.Combobox(self.JanelaCC, textvariable=self.variable6)
        self.nome22["width"] = 9
        self.nome22.place(x=305, y=155)

        self.nome22['values'] = [es[m] for m in range(0, 3)]




        self.nome46Label = Label(self.JanelaCC, text="Se C2, profundidade (m):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome46Label.place(x=60, y=185)

        pf = ['Selecione', '0.5', '0.6', '0.7', '0.8', '0.9', '1']
        self.variable_pf = tk.StringVar(self.JanelaCC)
        self.variable_pf.set(pf[0])

        self.nome46 = ttk.Combobox(self.JanelaCC, textvariable=self.variable_pf)
        self.nome46["width"] = 9
        self.nome46.place(x=305, y=185)

        self.nome46['values'] = [pf[m] for m in range(0, 7)]




        self.nome47Label = Label(self.JanelaCC, text="Se C3 ou C4, temperatura (°C):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome47Label.place(x=60, y=215)

        tp = ['Selecione', '20', '30', '40']

        self.variable_tp = tk.StringVar(self.JanelaCC)
        self.variable_tp.set(tp[0])

        self.nome47 = ttk.Combobox(self.JanelaCC, textvariable=self.variable_tp)
        self.nome47["width"] = 9
        self.nome47.place(x=305, y=215)

        self.nome47['values'] = [tp[m] for m in range(0, 4)]



        self.calcular4 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.calcular4["text"] = "Calcular"
        self.calcular4["font"] = ("Calibri", "8", "bold")
        self.calcular4["width"] = 12
        self.calcular4["command"] = self.Cap_cond_coor
        self.calcular4.place(x=90, y=245)

        self.mensagem4 = Label(self.JanelaCC, text="", font=self.fontePadrao, foreground='deep sky blue', bg='gray18')
        self.mensagem4.place(x=20, y=275)

        self.conf_sec = Label(self.JanelaCC, text="Confirme a seção do condutor:", font=("Arial", 10, "bold", "italic"), fg='white', bg='gray18')
        self.conf_sec.place(x=15, y=305)

        sec2 = ["Selecione", 1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240, 300, 400]

        self.variable_sec = tk.StringVar(self.JanelaCC)
        self.variable_sec.set(sec2[0])

        self.conf_sec = ttk.Combobox(self.JanelaCC, textvariable=self.variable_sec)
        self.conf_sec["width"] = 9
        self.conf_sec.place(x=212, y=305)

        self.conf_sec['values'] = [sec2[m] for m in range(3, 18)]







        # QUEDA DE TENSÃO
        # ADOTAR SEÇÃO MINIMA DE 4mm²


        self.titulo = Label(self.JanelaCC, text="Queda de tensão", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=530, y=5)

        #self.nome16Label = Label(self.JanelaCC, text="Potência módulo:", font=self.fontePadrao)
        #self.nome16Label.place(x=430, y=35)

        #self.nome16 = Entry(self.JanelaCC)
        #self.nome16["width"] = 10
        #self.nome16["font"] = self.fontePadrao
        #self.nome16.place(x=537, y=35)

        self.nome18Label = Label(self.JanelaCC, text="Tensão de operação do módulo (NMOT):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome18Label.place(x=430, y=35)

        self.nome18 = Entry(self.JanelaCC)
        self.nome18["width"] = 10
        self.nome18["font"] = self.fontePadrao
        self.nome18.place(x=668, y=35)

        self.nome19Label = Label(self.JanelaCC, text="N° modulos em série:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome19Label.place(x=430, y=65)

        self.nome19 = Entry(self.JanelaCC)
        self.nome19["width"] = 10
        self.nome19["font"] = self.fontePadrao
        self.nome19.place(x=668, y=65)

        self.nome24Label = Label(self.JanelaCC, text="Queda de tensão admitida em %:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome24Label.place(x=430, y=95)

        self.nome24 = Entry(self.JanelaCC)
        self.nome24["width"] = 10
        self.nome24["font"] = self.fontePadrao
        self.nome24.place(x=668, y=95)

        self.nome25Label = Label(self.JanelaCC, text="Comprimento do circuito em m:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome25Label.place(x=430, y=125)

        self.nome25 = Entry(self.JanelaCC)
        self.nome25["width"] = 10
        self.nome25["font"] = self.fontePadrao
        self.nome25.place(x=668, y=125)

        self.calcular5 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.calcular5["text"] = "Calcular"
        self.calcular5["font"] = ("Calibri", "8", "bold")
        self.calcular5["width"] = 12
        self.calcular5["command"] = self.Queda_tensao
        self.calcular5.place(x=500, y=155)

        self.mensagem5 = Label(self.JanelaCC, text="", font=self.fontePadrao, foreground='deep sky blue', bg='gray18')
        self.mensagem5.place(x=405, y=185)





        # PROTEÇÃO CONTRA SOBRECORRENTE DO MÓDULO FOTOVOLTAICO


        self.titulo = Label(self.JanelaCC, text="Proteção contra sobrecorrente", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=150, y=345)

        self.nome26Label = Label(self.JanelaCC, text="Informe a corrente de curto circuito na condição STC do módulo fotovoltaico:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome26Label.place(x=15, y=375)

        self.nome26 = Entry(self.JanelaCC)
        self.nome26["width"] = 10
        self.nome26["font"] = self.fontePadrao
        self.nome26.place(x=460, y=375)

        self.nome27Label = Label(self.JanelaCC,text="Informe o limite de corrente reversa do módulo fotovoltaico (Imodmáx,ocpr):",font=self.fontePadrao, fg='white', bg='gray18')
        self.nome27Label.place(x=15, y=405)

        self.nome27 = Entry(self.JanelaCC)
        self.nome27["width"] = 10
        self.nome27["font"] = self.fontePadrao
        self.nome27.place(x=460, y=405)

        self.nome28Label = Label(self.JanelaCC, text="N° máximo de séries conectadas em paralelo:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome28Label.place(x=15, y=435)

        self.nome28 = Entry(self.JanelaCC)
        self.nome28["width"] = 10
        self.nome28["font"] = self.fontePadrao
        self.nome28.place(x=460, y=435)

        #self.nome28Label = Label(self.JanelaCC,text="Informe o número total de séries fotovoltaicas conectadas em paralelo:",font=self.fontePadrao)
        #self.nome28Label.place(x=15, y=435)
#
        #self.nome28 = Entry(self.JanelaCC)
        #self.nome28["width"] = 10
        #self.nome28["font"] = self.fontePadrao
        #self.nome28.place(x=427, y=435)

        self.nome29Label = Label(self.JanelaCC,text="Informe o número total de séries fotovoltaicas do sistema:",font=("Arial", 10, "bold", "italic"), fg='white', bg='gray18')
        self.nome29Label.place(x=15, y=465)

        self.nome29 = Entry(self.JanelaCC)
        self.nome29["width"] = 10
        self.nome29["font"] = self.fontePadrao
        self.nome29.place(x=460, y=465)

        self.calcular6 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.calcular6["text"] = "Calcular"
        self.calcular6["font"] = ("Calibri", "8", "bold")
        self.calcular6["width"] = 12
        self.calcular6["command"] = self.Prot_sobrecorrente
        self.calcular6.place(x=200, y=495)

        self.mensagem6 = Label(self.JanelaCC, text="", font=self.fontePadrao, foreground='deep sky blue', bg='gray18')
        self.mensagem6.place(x=80, y=525)





        # JANELAS COM AS TEBLAS DA NORMA


        self.janela8 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.janela8["text"] = "Metodo C1"
        self.janela8["font"] = ("Calibri", "8", "bold")
        self.janela8["width"] = 30
        self.janela8["command"] = self.Metodo_C1
        self.janela8.place(x=570, y=362)

        self.janela9 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.janela9["text"] = "Metodo C2"
        self.janela9["font"] = ("Calibri", "8", "bold")
        self.janela9["width"] = 30
        self.janela9["command"] = self.Metodo_C2
        self.janela9.place(x=570, y=385)

        self.janela10 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.janela10["text"] = "Metodo C3"
        self.janela10["font"] = ("Calibri", "8", "bold")
        self.janela10["width"] = 30
        self.janela10["command"] = self.Metodo_C3
        self.janela10.place(x=570, y=408)

        self.janela11 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.janela11["text"] = "Metodo C4"
        self.janela11["font"] = ("Calibri", "8", "bold")
        self.janela11["width"] = 30
        self.janela11["command"] = self.Metodo_C4
        self.janela11.place(x=570, y=431)

        self.janela12 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.janela12["text"] = "Fator correção agrup. em feixe"
        self.janela12["font"] = ("Calibri", "8", "bold")
        self.janela12["width"] = 30
        self.janela12["command"] = self.fat_corr_agrup_feixe
        self.janela12.place(x=570, y=454)

        self.janela13 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.janela13["text"] = "Fator correção agrup. elet. enterrado"
        self.janela13["font"] = ("Calibri", "8", "bold")
        self.janela13["width"] = 30
        self.janela13["command"] = self.fat_corr_agrup_elet_enterrado
        self.janela13.place(x=570, y=477)

        self.gerar1 = Button(self.JanelaCC, fg='darkorange2', bg='white')
        self.gerar1["text"] = "Gerar lista"
        self.gerar1["font"] = ("Calibri", "8", "bold")
        self.gerar1["width"] = 30
        self.gerar1["command"] = self.lista_materiais_CC
        self.gerar1.place(x=570, y=500)





    def Cap_cond_coor(self):

        sec1 = [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240, 300, 400]

        imp = float(self.nome17.get().replace(",", "."))
        fa = float(self.nome20.get().replace(",", "."))
        met = self.variable4.get().upper()
        es = self.variable6.get().upper()
        mod = self.variable5.get()
        pf = self.variable_pf.get().replace(",", ".")
        tp = self.variable_tp.get()

        ic = imp / fa


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 1, PROTEGIDO DO SOL
        if met == 'C1' and mod == '1' and es == 'N' and ic > 0 and ic <= 24:
           self.mensagem4["font"] = ("Arial", "10", "bold")
           self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 24 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 32 and ic <= 42:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 42 and ic <= 53:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 53 and ic <= 74:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 74 and ic <= 98:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 98 and ic <= 131:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 131 and ic <= 163:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 163 and ic <= 205:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 205 and ic <= 255:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 255 and ic <= 307:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 307 and ic <= 360:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 360 and ic <= 415:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 415 and ic <= 474:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 474 and ic <= 568:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 568 and ic <= 656:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '1' and es == 'N' and ic > 656 and ic <= 790:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"



        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 2, PROTEGIDO DO SOL
        if met == 'C1' and mod == '2' and es == 'N' and ic > 0 and ic <= 23:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 23 and ic <= 31:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 31 and ic <= 41:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 41 and ic <= 53:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 53 and ic <= 74:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 74 and ic <= 98:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 98 and ic <= 131:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 131 and ic <= 164:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 164 and ic <= 208:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 208 and ic <= 259:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 259 and ic <= 313:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 367 and ic <= 367:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 367 and ic <= 424:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 424 and ic <= 484:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 484 and ic <= 581:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 581 and ic <= 671:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '2' and es == 'N' and ic > 671 and ic <= 808:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"



        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 3, PROTEGIDO DO SOL
        if met == 'C1' and mod == '3' and es == 'N' and ic > 0 and ic <= 27:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 27 and ic <= 36:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 36 and ic <= 48:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 48 and ic <= 61:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 61 and ic <= 85:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 85 and ic <= 112:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 112 and ic <= 149:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 149 and ic <= 185:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 185 and ic <= 233:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 233 and ic <= 291:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 291 and ic <= 350:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 350 and ic <= 411:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 411 and ic <= 473:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 473 and ic <= 539:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 539 and ic <= 645:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 645 and ic <= 746:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '3' and es == 'N' and ic > 746 and ic <= 897:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"



        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 4, PROTEGIDO DO SOL
        if met == 'C1' and mod == '4' and es == 'N' and ic > 0 and ic <= 23:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 23 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 32 and ic <= 42:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 42 and ic <= 54:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 54 and ic <= 76:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 76 and ic <= 101:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 101 and ic <= 136:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 136 and ic <= 170:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 170 and ic <= 215:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 215 and ic <= 270:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 270 and ic <= 326:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 326 and ic <= 384:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 384 and ic <= 444:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 444 and ic <= 508:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 508 and ic <= 611:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 611 and ic <= 708:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '4' and es == 'N' and ic > 708 and ic <= 854:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"







        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 1, EXPOSTO AO SOL
        if met == 'C1' and mod == '1' and es == 'S' and ic > 0 and ic <= 20:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 20 and ic <= 26:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 26 and ic <= 35:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 35 and ic <= 44:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 44 and ic <= 61:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 61 and ic <= 79:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 79 and ic <= 104:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 104 and ic <= 128:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 128 and ic <= 159:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 159 and ic <= 196:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 196 and ic <= 233:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 233 and ic <= 271:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 271 and ic <= 308:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 308 and ic <= 347:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 347 and ic <= 411:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 411 and ic <= 471:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '1' and es == 'S' and ic > 471 and ic <= 558:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"
#
        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 2, EXPOSTO AO SOL
        if met == 'C1' and mod == '2' and es == 'S' and ic > 0 and ic <= 19:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 19 and ic <= 26:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 26 and ic <= 34:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 34 and ic <= 43:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 43 and ic <= 60:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 60 and ic <= 79:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 79 and ic <= 105:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 105 and ic <= 130:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 130 and ic <= 163:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 163 and ic <= 201:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 201 and ic <= 241:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 241 and ic <= 281:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 281 and ic <= 321:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 321 and ic <= 363:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 363 and ic <= 431:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 431 and ic <= 494:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '2' and es == 'S' and ic > 494 and ic <= 586:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"
#
#
        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 3, EXPOSTO AO SOL
        if met == 'C1' and mod == '3' and es == 'S' and ic > 0 and ic <= 24:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 24 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 32 and ic <= 42:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 42 and ic <= 53:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 53 and ic <= 74:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 74 and ic <= 97:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 97 and ic <= 127:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 127 and ic <= 157:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 157 and ic <= 197:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 197 and ic <= 244:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 244 and ic <= 291:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 291 and ic <= 340:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 340 and ic <= 388:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 388 and ic <= 439:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 439 and ic <= 523:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 523 and ic <= 601:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '3' and es == 'S' and ic > 601 and ic <= 716:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"



        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C1, MODO DE INSTALAÇÃO 4, EXPOSTO AO SOL
        if met == 'C1' and mod == '4' and es == 'S' and ic > 0 and ic <= 20:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 20 and ic <= 26:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 26 and ic <= 35:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 35 and ic <= 45:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 45 and ic <= 62:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 62 and ic <= 83:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 83 and ic <= 110:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 110 and ic <= 137:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 137 and ic <= 173:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 173 and ic <= 216:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 216 and ic <= 259:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 259 and ic <= 304:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 304 and ic <= 349:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 349 and ic <= 397:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 397 and ic <= 475:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 475 and ic <= 547:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C1' and mod == '4' and es == 'S' and ic > 547 and ic <= 656:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"







        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C2, PROFUNDIDADE 0,5 m
        if met == 'C2' and pf == '0.5' and ic > 0 and ic <= 25:
             self.mensagem4["font"] = ("Arial", "10", "bold")
             self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C2' and pf == '0.5' and ic > 25 and ic <= 33:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C2' and pf == '0.5' and ic > 33 and ic <= 42:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C2' and pf == '0.5' and ic > 42 and ic <= 53:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C2' and pf == '0.5' and ic > 53 and ic <= 71:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C2' and pf == '0.5' and ic > 71 and ic <= 91:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C2' and pf == '0.5' and ic > 91 and ic <= 115:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C2' and pf == '0.5' and ic > 115 and ic <= 139:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C2' and pf == '0.5' and ic > 139 and ic <= 170:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C2' and pf == '0.5' and ic > 170 and ic <= 205:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C2' and pf == '0.5' and ic > 205 and ic <= 240:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C2' and pf == '0.5' and  ic > 240 and ic <= 274:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C2' and pf == '0.5' and ic > 274 and ic <= 309:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C2' and pf == '0.5' and ic > 309 and ic <= 346:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C2' and pf == '0.5' and ic > 346 and ic <= 404:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C2' and pf == '0.5' and ic > 404 and ic <= 457:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C2' and pf == '0.5' and ic > 457 and ic <= 534:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C2, PROFUNDIDADE 0,6 m
        if met == 'C2' and pf == '0.6' and ic > 0 and ic <= 25:
             self.mensagem4["font"] = ("Arial", "10", "bold")
             self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C2' and pf == '0.6' and ic > 25 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C2' and pf == '0.6' and ic > 32 and ic <= 42:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C2' and pf == '0.6' and ic > 42 and ic <= 52:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C2' and pf == '0.6' and ic > 52 and ic <= 70:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C2' and pf == '0.6' and ic > 70 and ic <= 89:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C2' and pf == '0.6' and ic > 89 and ic <= 114:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C2' and pf == '0.6' and ic > 114 and ic <= 137:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C2' and pf == '0.6' and ic > 137 and ic <= 167:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C2' and pf == '0.6' and ic > 167 and ic <= 202:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C2' and pf == '0.6' and ic > 202 and ic <= 235:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C2' and pf == '0.6' and  ic > 235 and ic <= 269:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C2' and pf == '0.6' and ic > 269 and ic <= 303:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C2' and pf == '0.6' and ic > 303 and ic <= 339:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C2' and pf == '0.6' and ic > 339 and ic <= 396:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C2' and pf == '0.6' and ic > 396 and ic <= 447:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C2' and pf == '0.6' and ic > 447 and ic <= 523:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C2, PROFUNDIDADE 0,7 m
        if met == 'C2' and pf == '0.7' and ic > 0 and ic <= 24:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C2' and pf == '0.7' and ic > 24 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C2' and pf == '0.7' and ic > 32 and ic <= 41:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C2' and pf == '0.7' and ic > 41 and ic <= 51:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C2' and pf == '0.7' and ic > 51 and ic <= 69:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C2' and pf == '0.7' and ic > 69 and ic <= 88:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C2' and pf == '0.7' and ic > 88 and ic <= 112:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C2' and pf == '0.7' and ic > 112 and ic <= 135:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C2' and pf == '0.7' and ic > 135 and ic <= 164:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C2' and pf == '0.7' and ic > 164 and ic <= 199:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C2' and pf == '0.7' and ic > 199 and ic <= 232:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C2' and pf == '0.7' and ic > 232 and ic <= 265:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C2' and pf == '0.7' and ic > 265 and ic <= 299:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C2' and pf == '0.7' and ic > 299 and ic <= 334:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C2' and pf == '0.7' and ic > 334 and ic <= 389:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C2' and pf == '0.7' and ic > 389 and ic <= 440:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C2' and pf == '0.7' and ic > 440 and ic <= 514:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C2, PROFUNDIDADE 0,8 m
        if met == 'C2' and pf == '0.8' and ic > 0 and ic <= 24:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C2' and pf == '0.8' and ic > 24 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C2' and pf == '0.8' and ic > 32 and ic <= 41:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C2' and pf == '0.8' and ic > 41 and ic <= 51:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C2' and pf == '0.8' and ic > 51 and ic <= 68:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C2' and pf == '0.8' and ic > 68 and ic <= 87:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C2' and pf == '0.8' and ic > 87 and ic <= 111:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C2' and pf == '0.8' and ic > 111 and ic <= 133:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C2' and pf == '0.8' and ic > 133 and ic <= 163:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C2' and pf == '0.8' and ic > 163 and ic <= 196:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C2' and pf == '0.8' and ic > 196 and ic <= 229:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C2' and pf == '0.8' and ic > 229 and ic <= 262:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C2' and pf == '0.8' and ic > 262 and ic <= 295:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C2' and pf == '0.8' and ic > 295 and ic <= 329:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C2' and pf == '0.8' and ic > 329 and ic <= 384:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C2' and pf == '0.8' and ic > 384 and ic <= 434:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C2' and pf == '0.8' and ic > 434 and ic <= 506:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C2, PROFUNDIDADE 0,8 m
        if met == 'C2' and pf == '0.9' and ic > 0 and ic <= 24:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C2' and pf == '0.9' and ic > 24 and ic <= 31:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C2' and pf == '0.9' and ic > 31 and ic <= 41:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C2' and pf == '0.9' and ic > 41 and ic <= 50:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C2' and pf == '0.9' and ic > 50 and ic <= 68:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C2' and pf == '0.9' and ic > 68 and ic <= 86:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C2' and pf == '0.9' and ic > 86 and ic <= 110:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C2' and pf == '0.9' and ic > 110 and ic <= 132:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C2' and pf == '0.9' and ic > 132 and ic <= 161:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C2' and pf == '0.9' and ic > 161 and ic <= 194:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C2' and pf == '0.9' and ic > 194 and ic <= 226:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C2' and pf == '0.9' and ic > 226 and ic <= 259:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C2' and pf == '0.9' and ic > 259 and ic <= 292:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C2' and pf == '0.9' and ic > 292 and ic <= 326:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C2' and pf == '0.9' and ic > 326 and ic <= 379:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C2' and pf == '0.9' and ic > 379 and ic <= 429:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C2' and pf == '0.9' and ic > 429 and ic <= 500:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C2, PROFUNDIDADE 0,8 m
        if met == 'C2' and pf == '1' and ic > 0 and ic <= 24:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C2' and pf == '1' and ic > 24 and ic <= 31:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C2' and pf == '1' and ic > 31 and ic <= 40:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C2' and pf == '1' and ic > 40 and ic <= 50:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C2' and pf == '1' and ic > 50 and ic <= 67:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C2' and pf == '1' and ic > 67 and ic <= 86:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C2' and pf == '1' and ic > 86 and ic <= 109:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C2' and pf == '1' and ic > 109 and ic <= 131:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C2' and pf == '1' and ic > 131 and ic <= 159:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C2' and pf == '1' and ic > 159 and ic <= 193:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C2' and pf == '1' and ic > 193 and ic <= 224:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C2' and pf == '1' and ic > 224 and ic <= 256:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C2' and pf == '1' and ic > 256 and ic <= 289:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C2' and pf == '1' and ic > 289 and ic <= 322:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C2' and pf == '1' and ic > 322 and ic <= 375:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C2' and pf == '1' and ic > 375 and ic <= 424:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C2' and pf == '1' and ic > 424 and ic <= 495:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"






        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C3, TEMPERATURA ATÉ 20°
        if met == 'C3' and tp == '20' and ic > 0 and ic <= 22:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C3' and tp == '20' and ic > 22 and ic <= 29:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C3' and tp == '20' and ic > 29 and ic <= 37:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C3' and tp == '20' and ic > 37 and ic <= 46:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C3' and tp == '20' and ic > 46 and ic <= 62:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C3' and tp == '20' and ic > 62 and ic <= 79:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C3' and tp == '20' and ic > 79 and ic <= 102:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C3' and tp == '20' and ic > 102 and ic <= 124:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C3' and tp == '20' and ic > 124 and ic <= 151:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C3' and tp == '20' and ic > 151 and ic <= 186:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C3' and tp == '20' and ic > 186 and ic <= 217:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C3' and tp == '20' and ic > 217 and ic <= 250:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C3' and tp == '20' and ic > 250 and ic <= 287:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C3' and tp == '20' and ic > 287 and ic <= 321:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C3' and tp == '20' and ic > 321 and ic <= 380:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C3' and tp == '20' and ic > 380 and ic <= 429:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C3' and tp == '20' and ic > 429 and ic <= 503:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C3, TEMPERATURA ATÉ 30°
        if met == 'C3' and tp == '30' and ic > 0 and ic <= 20:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C3' and tp == '30' and ic > 20 and ic <= 27:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C3' and tp == '30' and ic > 27 and ic <= 34:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C3' and tp == '30' and ic > 34 and ic <= 42:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C3' and tp == '30' and ic > 42 and ic <= 58:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C3' and tp == '30' and ic > 58 and ic <= 74:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C3' and tp == '30' and ic > 74 and ic <= 94:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C3' and tp == '30' and ic > 94 and ic <= 115:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C3' and tp == '30' and ic > 115 and ic <= 140:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C3' and tp == '30' and ic > 140 and ic <= 172:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C3' and tp == '30' and ic > 172 and ic <= 201:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C3' and tp == '30' and ic > 201 and ic <= 232:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C3' and tp == '30' and ic > 232 and ic <= 266:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C3' and tp == '30' and ic > 266 and ic <= 297:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C3' and tp == '30' and ic > 297 and ic <= 352:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C3' and tp == '30' and ic > 352 and ic <= 397:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C3' and tp == '30' and ic > 397 and ic <= 466:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C3, TEMPERATURA ATÉ 40°
        if met == 'C3' and tp == '40' and ic > 0 and ic <= 19:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C3' and tp == '40' and ic > 19 and ic <= 24:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C3' and tp == '40' and ic > 24 and ic <= 31:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C3' and tp == '40' and ic > 31 and ic <= 39:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C3' and tp == '40' and ic > 39 and ic <= 53:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C3' and tp == '40' and ic > 53 and ic <= 67:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C3' and tp == '40' and ic > 67 and ic <= 86:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C3' and tp == '40' and ic > 86 and ic <= 105:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C3' and tp == '40' and ic > 105 and ic <= 128:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"
        elif met == 'C3' and tp == '40' and ic > 128 and ic <= 157:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[9]}"
        elif met == 'C3' and tp == '40' and ic > 157 and ic <= 183:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[10]}"
        elif met == 'C3' and tp == '40' and ic > 183 and ic <= 212:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[11]}"
        elif met == 'C3' and tp == '40' and ic > 212 and ic <= 243:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[12]}"
        elif met == 'C3' and tp == '40' and ic > 243 and ic <= 271:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[13]}"
        elif met == 'C3' and tp == '40' and ic > 271 and ic <= 321:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[14]}"
        elif met == 'C3' and tp == '40' and ic > 321 and ic <= 362:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[15]}"
        elif met == 'C3' and tp == '40' and ic > 362 and ic <= 425:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[16]}"






        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C4, TEMPERATURA ATÉ 20°
        if met == 'C4' and tp == '20' and ic > 0 and ic <= 25:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C4' and tp == '20' and ic > 25 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C4' and tp == '20' and ic > 32 and ic <= 42:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C4' and tp == '20' and ic > 42 and ic <= 52:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C4' and tp == '20' and ic > 52 and ic <= 73:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C4' and tp == '20' and ic > 73 and ic <= 93:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C4' and tp == '20' and ic > 93 and ic <= 121:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C4' and tp == '20' and ic > 121 and ic <= 150:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C4' and tp == '20' and ic > 150 and ic <= 184:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C4, TEMPERATURA ATÉ 30°
        if met == 'C4' and tp == '30' and ic > 0 and ic <= 22:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C4' and tp == '30' and ic > 22 and ic <= 29:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C4' and tp == '30' and ic > 29 and ic <= 37:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C4' and tp == '30' and ic > 37 and ic <= 46:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C4' and tp == '30' and ic > 46 and ic <= 64:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C4' and tp == '30' and ic > 64 and ic <= 83:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C4' and tp == '30' and ic > 83 and ic <= 107:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C4' and tp == '30' and ic > 107 and ic <= 133:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C4' and tp == '30' and ic > 133 and ic <= 163:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"


        # DIMENSIONAMENTE PARA MÉTODO DE INSTALAÇÃO C4, TEMPERATURA ATÉ 40°
        if met == 'C4' and tp == '40' and ic > 0 and ic <= 19:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[0]}"
        elif met == 'C4' and tp == '40' and ic > 19 and ic <= 24:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[1]}"
        elif met == 'C4' and tp == '40' and ic > 24 and ic <= 32:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[2]}"
        elif met == 'C4' and tp == '40' and ic > 32 and ic <= 39:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[3]}"
        elif met == 'C4' and tp == '40' and ic > 39 and ic <= 55:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[4]}"
        elif met == 'C4' and tp == '40' and ic > 55 and ic <= 71:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[5]}"
        elif met == 'C4' and tp == '40' and ic > 71 and ic <= 92:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[6]}"
        elif met == 'C4' and tp == '40' and ic > 92 and ic <= 114:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[7]}"
        elif met == 'C4' and tp == '40' and ic > 114 and ic <= 140:
            self.mensagem4["font"] = ("Arial", "10", "bold")
            self.mensagem4["text"] = f"A seção do condutor deverá ser em mm²: {sec1[8]}"



    def Queda_tensao(self):

        # DIMENSIONAMENTO FV CONFORME A NBR 16612
        # QUEDA DE TENSÃO
        # ADOTAR SEÇÃO MINIMA DE 4mm²

        #p = float(self.nome16.get().replace(",", "."))
        imp = float(self.nome17.get().replace(",", "."))
        vmp = float(self.nome18.get().replace(",", "."))
        ns = int(self.nome19.get())
        eperc = float(self.nome24.get().replace(",", "."))
        l1 = float(self.nome25.get().replace(",", "."))

        cond = 45.5
        vtot = vmp * ns

        e = vtot * (eperc / 100)
        #epercmax = (e / vtot) * 100

        secao = (l1 * 2 * imp) / (cond * e)

        self.mensagem5["font"] = ("Arial", "10", "bold")
        self.mensagem5["text"] = f"A seção nominal do condutor exigida é de {secao:.1f}"


    def Prot_sobrecorrente(self):

        # DIMENSIONAMENTO FV CONFORME A NBR 16612
        # PROTEÇÃO CONTRA SOBRECORRENTE DO MÓDULO FOTOVOLTAICO

        isc = float(self.nome26.get().replace(",", "."))
        ici = float(self.nome27.get().replace(",", "."))
        sp = int(self.nome28.get())

        pcs = (sp - 1) * isc

        if pcs < ici:
            self.mensagem6["foreground"] = "deep sky blue"
            self.mensagem6["font"] = ("Arial", "10", "bold")
            self.mensagem6["text"] = f"Não necessita proteção contra sobrecorrente!"
        else:
            self.mensagem6["foreground"] = "red"
            self.mensagem6["font"] = ("Arial", "10", "bold")
            self.mensagem6["text"] = f"Deve ser fonercido proteção contra sobrecorrente!"

            self.inom = 1.5 * isc
            self.inom1 = 2.4 * isc

            if self.inom1 >= ici:
                self.mensagem6["foreground"] = "red"
                self.mensagem6["font"] = ("Arial", "10", "bold")
                self.mensagem6["text"] = f"""Deve ser fornecido proteção contra sobrecorrente!\nEspecificar fusível na faixa {self.inom:.2f} < in < {ici}"""
            else:
                self.mensagem6["foreground"] = "red"
                self.mensagem6["font"] = ("Arial", "10", "bold")
                self.mensagem6["text"] = f"""Deve ser fornecido proteção contra sobrecorrente!\nEspecificar fusível na faixa {self.inom:.2f} < in < {self.inom1:.2f}"""


    def Metodo_C1(self):

        self.janela8 = Toplevel()
        self.janela8.title("Capacidade de condução de corrente – Método C.1")
        self.janela8.iconbitmap('baixados.ico')
        self.imag6 = PhotoImage(file="met_c1.png")
        self.figura6 = Label(self.janela8, image=self.imag6)
        self.figura6.pack()

    def Metodo_C2(self):

        self.janela9 = Toplevel()
        self.janela9.title("Capacidade de condução de corrente – Método C.2")
        self.janela9.iconbitmap('baixados.ico')
        self.imag7 = PhotoImage(file="met_c2.png")
        self.figura7 = Label(self.janela9, image=self.imag7)
        self.figura7.pack()

    def Metodo_C3(self):

        self.janela10 = Toplevel()
        self.janela10.title("Capacidade de condução de corrente – Método C.3")
        self.janela10.iconbitmap('baixados.ico')
        self.imag8 = PhotoImage(file="met_c3.png")
        self.figura8 = Label(self.janela10, image=self.imag8)
        self.figura8.pack()

    def Metodo_C4(self):

        self.janela11 = Toplevel()
        self.janela11.title("Capacidade de condução de corrente – Método C.4")
        self.janela11.iconbitmap('baixados.ico')
        self.imag9 = PhotoImage(file="met_c4.png")
        self.figura9 = Label(self.janela11, image=self.imag9)
        self.figura9.pack()

    def fat_corr_agrup_feixe(self):

        self.janela12 = Toplevel()
        self.janela12.title("Fator de correção aplicáveis em condutores agrupados em feixe")
        self.janela12.iconbitmap('baixados.ico')
        self.imag10 = PhotoImage(file="fat_corr_agrup_feixe.png")
        self.figura10 = Label(self.janela12, image=self.imag10)
        self.figura10.pack()

    def fat_corr_agrup_elet_enterrado(self):

        self.janela13 = Toplevel()
        self.janela13.title("Fator de agrupamento para linhas em eletrodutos enterrados")
        self.janela13.iconbitmap('baixados.ico')
        self.imag11 = PhotoImage(file="fat_corr_agrup_elet_enterrado.png")
        self.figura11 = Label(self.janela13, image=self.imag11)
        self.figura11.pack()

    def lista_materiais_CC(self):

        # GERA LISTA DE MATERIAIS CC

        #ncc = int(self.nome6.get())
        #l = float(self.nome10.get().replace(",", "."))
        #sec = float(self.nome13.get().replace(",", "."))
        #dsj = float(self.variable.get())
       # te = int(self.variable1.get())


        l1 = float(self.nome25.get().replace(",", "."))
        total_series_fv = int(self.nome29.get())
        secCC = float(self.variable_sec.get().replace(",", "."))

        self.gerar1 = xlsxwriter.Workbook('Lista de Materiais CC.xlsx')
        self.worksheet = self.gerar1.add_worksheet('Lista')

        dados1 = [
            ["Produto", "Característica", "Quantidade"],
            ["Condutor CC/FV Vermelho", f"{secCC} mm²", f"{(l1 * total_series_fv) + (0.1*(l1 * total_series_fv))} m"],
            ["Condutor CC/FV Preto", f"{secCC} mm²", f"{(l1 * total_series_fv) + (0.1*(l1 * total_series_fv))} m"],
            ["Conectores MC4", "Par - Macho + Fêmea", f"{total_series_fv}"]
            #["Eletroduto", f"{te} mm", f"{(l * 1000)+(0.1*(l * 1000))} m"],
        ]
        row = 0
        col = 0
        for prod1, secao1, qtde1 in (dados1):
            self.worksheet.write(row, col, prod1)
            self.worksheet.write(row, col + 1, secao1)
            self.worksheet.write(row, col + 2, qtde1)
            row += 1

        self.gerar1.close()










    def Janela_cadastro(self):

        self.Janela_cadastro = Toplevel()
        self.Janela_cadastro.geometry("790x480")
        self.Janela_cadastro.minsize(790, 480)
        self.Janela_cadastro.maxsize(790, 480)
        self.Janela_cadastro.configure(bg='gray18')
        self.Janela_cadastro.title("Cadastro Equipamentos")
        self.Janela_cadastro.iconbitmap('baixados.ico')

        self.fontePadrao = ("Arial", "10")



        # MÓDULO FV


        self.titulo = Label(self.Janela_cadastro, text="Dados módulo fotovoltaico", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=110, y=5)

        self.fabricanteLabel = Label(self.Janela_cadastro, text="Fabricante:", font=self.fontePadrao, fg='white', bg='gray18')
        self.fabricanteLabel.place(x=40, y=35)

        self.fabricante = Entry(self.Janela_cadastro)
        self.fabricante["width"] = 20
        self.fabricante["font"] = self.fontePadrao
        self.fabricante.place(x=271, y=35)

        self.modeloLabel = Label(self.Janela_cadastro, text="Modelo:", font=self.fontePadrao, fg='white', bg='gray18')
        self.modeloLabel.place(x=40, y=65)

        self.modelo = Entry(self.Janela_cadastro)
        self.modelo["width"] = 20
        self.modelo["font"] = self.fontePadrao
        self.modelo.place(x=271, y=65)

        self.nome30Label = Label(self.Janela_cadastro, text="Potência do módulo (STC):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome30Label.place(x=40, y=95)

        self.nome30 = Entry(self.Janela_cadastro)
        self.nome30["width"] = 20
        self.nome30["font"] = self.fontePadrao
        self.nome30.place(x=271, y=95)

        self.nome31Label = Label(self.Janela_cadastro, text="Tensão de circuito aberto (STC):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome31Label.place(x=40, y=125)

        self.nome31 = Entry(self.Janela_cadastro)
        self.nome31["width"] = 20
        self.nome31["font"] = self.fontePadrao
        self.nome31.place(x=271, y=125)

        self.nome32Label = Label(self.Janela_cadastro, text="Corrente de curto circuito (STC):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome32Label.place(x=40, y=155)

        self.nome32 = Entry(self.Janela_cadastro)
        self.nome32["width"] = 20
        self.nome32["font"] = self.fontePadrao
        self.nome32.place(x=271, y=155)

        self.nome33Label = Label(self.Janela_cadastro, text="Tensão de operação (NMOT):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome33Label.place(x=40, y=185)

        self.nome33 = Entry(self.Janela_cadastro)
        self.nome33["width"] = 20
        self.nome33["font"] = self.fontePadrao
        self.nome33.place(x=271, y=185)

        self.nome34Label = Label(self.Janela_cadastro, text="Potência de operação (NMOT):", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome34Label.place(x=40, y=215)

        self.nome34 = Entry(self.Janela_cadastro)
        self.nome34["width"] = 20
        self.nome34["font"] = self.fontePadrao
        self.nome34.place(x=271, y=215)

        self.nome35Label = Label(self.Janela_cadastro, text="Coeficiente de temperatura para Voc:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome35Label.place(x=40, y=245)

        self.nome35 = Entry(self.Janela_cadastro)
        self.nome35["width"] = 20
        self.nome35["font"] = self.fontePadrao
        self.nome35.place(x=271, y=245)

        self.nome36Label = Label(self.Janela_cadastro, text="Coeficiente de temperatura para Isc:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome36Label.place(x=40, y=275)

        self.nome36 = Entry(self.Janela_cadastro)
        self.nome36["width"] = 20
        self.nome36["font"] = self.fontePadrao
        self.nome36.place(x=271, y=275)

        self.nome37Label = Label(self.Janela_cadastro, text="Coeficiente de temperatura para Pmax:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome37Label.place(x=40, y=305)

        self.nome37 = Entry(self.Janela_cadastro)
        self.nome37["width"] = 20
        self.nome37["font"] = self.fontePadrao
        self.nome37.place(x=271, y=305)

        self.comprimentoLabel = Label(self.Janela_cadastro, text="Comprimento do módulo:", font=self.fontePadrao, fg='white', bg='gray18')
        self.comprimentoLabel.place(x=40, y=335)

        self.comprimento = Entry(self.Janela_cadastro)
        self.comprimento["width"] = 20
        self.comprimento["font"] = self.fontePadrao
        self.comprimento.place(x=271, y=335)

        self.larguraLabel = Label(self.Janela_cadastro, text="Largura do módulo:", font=self.fontePadrao, fg='white', bg='gray18')
        self.larguraLabel.place(x=40, y=365)

        self.largura = Entry(self.Janela_cadastro)
        self.largura["width"] = 20
        self.largura["font"] = self.fontePadrao
        self.largura.place(x=271, y=365)

        self.gerar_mod = Button(self.Janela_cadastro, fg='darkorange2', bg='white')
        self.gerar_mod["text"] = "Cadastrar módulo"
        self.gerar_mod["font"] = ("Calibri", "9", "bold")
        self.gerar_mod["width"] = 20
        self.gerar_mod["command"] = self.Cadastro_mod
        self.gerar_mod.place(x=155, y=405)




        # DADOS INVERSOR

        self.titulo = Label(self.Janela_cadastro, text="Dados Inversor", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "11", "bold")
        self.titulo.place(x=525, y=5)

        self.fabricante1Label = Label(self.Janela_cadastro, text="Fabricante:", font=self.fontePadrao, fg='white', bg='gray18')
        self.fabricante1Label.place(x=430, y=35)

        self.fabricante1 = Entry(self.Janela_cadastro)
        self.fabricante1["width"] = 20
        self.fabricante1["font"] = self.fontePadrao
        self.fabricante1.place(x=591, y=35)

        self.modelo1Label = Label(self.Janela_cadastro, text="Modelo:", font=self.fontePadrao, fg='white', bg='gray18')
        self.modelo1Label.place(x=430, y=65)

        self.modelo1 = Entry(self.Janela_cadastro)
        self.modelo1["width"] = 20
        self.modelo1["font"] = self.fontePadrao
        self.modelo1.place(x=591, y=65)

        self.nome38Label = Label(self.Janela_cadastro, text="Potência:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome38Label.place(x=430, y=95)

        self.nome38 = Entry(self.Janela_cadastro)
        self.nome38["width"] = 20
        self.nome38["font"] = self.fontePadrao
        self.nome38.place(x=591, y=95)

        self.nome39Label = Label(self.Janela_cadastro, text="Maxima tensão CC:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome39Label.place(x=430, y=125)

        self.nome39 = Entry(self.Janela_cadastro)
        self.nome39["width"] = 20
        self.nome39["font"] = self.fontePadrao
        self.nome39.place(x=591, y=125)

        #self.nome40Label = Label(self.Janela_string, text="Máxima corrente CC:", font=self.fontePadrao)
        #self.nome40Label.place(x=430, y=95)

        #self.nome40 = Entry(self.Janela_string)
        #self.nome40["width"] = 10
        #self.nome40["font"] = self.fontePadrao
        #self.nome40.place(x=558, y=95)

        self.nome41Label = Label(self.Janela_cadastro, text="Máxima corrente SPMP:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome41Label.place(x=430, y=155)

        self.nome41 = Entry(self.Janela_cadastro)
        self.nome41["width"] = 20
        self.nome41["font"] = self.fontePadrao
        self.nome41.place(x=591, y=155)

        self.nome42Label = Label(self.Janela_cadastro, text="Tensão mínima de SPMP:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome42Label.place(x=430, y=185)

        self.nome42 = Entry(self.Janela_cadastro)
        self.nome42["width"] = 20
        self.nome42["font"] = self.fontePadrao
        self.nome42.place(x=591, y=185)

        self.nome43Label = Label(self.Janela_cadastro, text="Tensão máxima de SPMP:", font=self.fontePadrao, fg='white', bg='gray18')
        self.nome43Label.place(x=430, y=215)

        self.nome43 = Entry(self.Janela_cadastro)
        self.nome43["width"] = 20
        self.nome43["font"] = self.fontePadrao
        self.nome43.place(x=591, y=215)

        self.gerar_inv= Button(self.Janela_cadastro, fg='darkorange2', bg='white')
        self.gerar_inv["text"] = "Cadastrar inversor"
        self.gerar_inv["font"] = ("Calibri", "9", "bold")
        self.gerar_inv["width"] = 20
        self.gerar_inv["command"] = self.Cadastro_inv
        self.gerar_inv.place(x=525, y=255)


    #def Cadastro_equip(self):
#
    #    # Módulo:
#
    #    pmax_stc = float(self.nome30.get().replace(",", "."))
    #    voc_stc = float(self.nome31.get().replace(",", "."))
    #    isc_stc = float(self.nome32.get().replace(",", "."))
    #    vop_nmot = float(self.nome33.get().replace(",", "."))
    #    pop_nmot = float(self.nome34.get().replace(",", "."))
    #    coef_voc = float(self.nome35.get().replace(",", "."))
    #    coef_isc = float(self.nome36.get().replace(",", "."))
    #    coef_pmax = float(self.nome37.get().replace(",", "."))
    #    comp = float(self.comprimento.get().replace(",", "."))
    #    larg = float(self.largura.get().replace(",", "."))
#
    #    # Inversor:
#
    #    pmax_inv = float(self.nome38.get().replace(",", "."))
    #    vmax_cc = float(self.nome39.get().replace(",", "."))
    #    #imax_cc = float(self.nome40.get().replace(",", "."))
    #    imax_spmp = float(self.nome41.get().replace(",", "."))
    #    vmin_spmp = float(self.nome42.get().replace(",", "."))
    #    vmax_spmp = float(self.nome43.get().replace(",", "."))


    def Cadastro_mod(self):


       # GERA LISTA DE MÓDULOS

        wb_obj = op.load_workbook('dados_mod.xlsx')
        sheet_obj = wb_obj.active

        fabricante = self.fabricante.get()
        modelo = self.modelo.get()
        pmax_stc = float(self.nome30.get().replace(",", "."))
        voc_stc = float(self.nome31.get().replace(",", "."))
        isc_stc = float(self.nome32.get().replace(",", "."))
        vop_nmot = float(self.nome33.get().replace(",", "."))
        pop_nmot = float(self.nome34.get().replace(",", "."))
        coef_voc = float(self.nome35.get().replace(",", "."))
        coef_isc = float(self.nome36.get().replace(",", "."))
        coef_pmax = float(self.nome37.get().replace(",", "."))
        comp = float(self.comprimento.get().replace(",", "."))
        larg = float(self.largura.get().replace(",", "."))

        to_append = [fabricante ,modelo, pmax_stc, voc_stc, isc_stc, vop_nmot, pop_nmot, coef_voc, coef_isc, coef_pmax, comp, larg]

        sheet_obj.append(to_append)

        wb_obj.save('dados_mod.xlsx')


    def Cadastro_inv(self):

        # GERA LISTA DE MÓDULOS

        wb_obj1 = op.load_workbook('dados_inv.xlsx')
        sheet_obj1 = wb_obj1.active

        fabricante = self.fabricante1.get()
        modelo = self.modelo1.get()
        pmax_inv = float(self.nome38.get().replace(",", "."))
        vmax_cc = float(self.nome39.get().replace(",", "."))
        imax_spmp = float(self.nome41.get().replace(",", "."))
        vmin_spmp = float(self.nome42.get().replace(",", "."))
        vmax_spmp = float(self.nome43.get().replace(",", "."))

        to_append = [fabricante, modelo, pmax_inv, vmax_cc, imax_spmp, vmin_spmp, vmax_spmp]

        sheet_obj1.append(to_append)

        wb_obj1.save('dados_inv.xlsx')













    def Janela_string(self):

        self.Janela_string = Toplevel()
        self.Janela_string.geometry("760x450")
        self.Janela_string.minsize(760, 450)
        self.Janela_string.maxsize(760, 450)
        self.Janela_string.configure(bg='gray18')
        self.Janela_string.title("Configuração String")
        self.Janela_string.iconbitmap('baixados.ico')

        self.fontePadrao = ("Arial", "10")

        wb_obj = op.load_workbook('dados_mod.xlsx')
        sheet_obj = wb_obj.active

        wb_obj1 = op.load_workbook('dados_inv.xlsx')
        sheet_obj1 = wb_obj1.active

        # Configuração

        self.titulo = Label(self.Janela_string, text="Configuração de String", bg='gray18', fg='orange')
        self.titulo["font"] = ("Arial", "12", "bold")
        self.titulo.place(x=300, y=20)

        self.nome46Label = Label(self.Janela_string, text="N° módulos em série:", font=self.fontePadrao, bg='gray18', fg='white')
        self.nome46Label.place(x=300, y=110)

        self.nome46 = Entry(self.Janela_string)
        self.nome46["width"] = 10
        self.nome46["font"] = self.fontePadrao
        self.nome46.place(x=433, y=110)

        self.nome47Label = Label(self.Janela_string, text="N° séries em paralelo:", font=self.fontePadrao, bg='gray18', fg='white')
        self.nome47Label.place(x=300, y=140)

        self.nome47 = Entry(self.Janela_string)
        self.nome47["width"] = 10
        self.nome47["font"] = self.fontePadrao
        self.nome47.place(x=433, y=140)



        self.modulo = Label(self.Janela_string, text="Escolha o módulo:",font=("Arial", 10, "bold", "italic"),  bg='gray18', fg='white')
        self.modulo.place(x=100, y=80)
        ##coluna = bd['Modelo']
        modulo = []
        ##inversor = bd['Modelo']

        max_row = sheet_obj.max_row
        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            modulo.append(cell_obj.value)

        self.variable_mod = tk.StringVar(self.Janela_string)
        self.variable_mod.set(modulo[0])


        self.modulo = ttk.Combobox(self.Janela_string, textvariable=self.variable_mod)
        self.modulo["width"] = 20
        self.modulo.place(x=222, y=80)


        self.modulo['values'] = [modulo[m] for m in range(len(modulo))]


        self.inversor = Label(self.Janela_string, text="Escolha o inversor:", font=("Arial", 10, "bold", "italic"),  bg='gray18', fg='white')
        self.inversor.place(x=380, y=80)
        ##coluna = bd['Modelo']
        inversor = []
        ##inversor = bd['Modelo']

        max_row1 = sheet_obj1.max_row
        for i in range(1, max_row1 + 1):
            cell_obj1 = sheet_obj1.cell(row=i, column=2)
            inversor.append(cell_obj1.value)

        self.variable_inv = tk.StringVar(self.Janela_string)
        self.variable_inv.set(inversor[0])

        self.inversor = ttk.Combobox(self.Janela_string, textvariable=self.variable_inv)
        self.inversor["width"] = 20
        self.inversor.place(x=504, y=80)

        self.inversor['values'] = [inversor[m] for m in range(len(inversor))]

        #self.index = str(inversor.current())

        self.calcular8 = Button(self.Janela_string,  bg='white', fg='darkorange2')
        self.calcular8["text"] = "Obter resultados"
        self.calcular8["font"] = ("Calibri", "11", "bold")
        self.calcular8["width"] = 30
        self.calcular8["command"] = self.dimensionamento_string
        self.calcular8.place(x=250, y=200)

        # self.textoContainer = Frame(self.calcular7)
        # self.textoContainer["padx"] = 20
        # self.textoContainer["pady"] = 5
        # self.textoContainer.place(x=20, y=350)



        self.mensagem8 = Label(self.Janela_string, text="", font=self.fontePadrao, bg='gray18')
        self.mensagem8.place(x=100, y=255)
        # self.mensagem8.pack(side=RIGHT)

    #def OptionCallBack(*args):
    #    index = modulo.current()
    #    index1 = inversor.current()

    def dimensionamento_string(self):

        # Módulo:

        wb_obj = op.load_workbook('dados_mod.xlsx')
        sheet_obj = wb_obj.active


        ind = int(self.modulo.current())


        dados1 = []
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        for j in range(3, max_col + 1):
            for k in range(2, max_row + 1):
                cell_obj = sheet_obj.cell(row=k, column=j)
                dados1.append(cell_obj.value)
                data = []
                for l in dados1[ind-1::max_row - 1]:
                    data.append(l)

        pmax_stc = data[0]
        voc_stc = data[1]
        isc_stc = data[2]
        vop_nmot = data[3]
        pop_nmot = data[4]
        coef_voc = data[5]
        coef_isc = data[6]
        coef_pmax = data[7]





        # Inversor:

        wb_obj1 = op.load_workbook('dados_inv.xlsx')
        sheet_obj1 = wb_obj1.active

        ind1 = int(self.inversor.current())

        dados2 = []
        max_col1 = sheet_obj1.max_column
        max_row1 = sheet_obj1.max_row
        for m in range(3, max_col1 + 1):
            for n in range(2, max_row1 + 1):
                cell_obj1 = sheet_obj1.cell(row=n, column=m)
                dados2.append(cell_obj1.value)
                data1 = []
                for o in dados2[ind1 - 1::max_row1 - 1]:
                    data1.append(o)

         #pmax_inv = self.bd.iat[16, 2]
         #vmax_cc = self.bd.iat[16, 3]
         ## imax_cc = float(self.nome40.get().replace(",", "."))
         #imax_spmp = self.bd.iat[16, 4]
         #vmin_spmp = self.bd.iat[16, 5]
         #vmax_spmp = self.bd.iat[16, 6]

        pmax_inv = data1[0]
        vmax_cc = data1[1]
        imax_spmp = data1[2]
        vmin_spmp = data1[3]
        vmax_spmp = data1[4]

        # Configuração

        ns_fv = int(self.nome46.get().replace(",", "."))
        np_fv = int(self.nome47.get().replace(",", "."))

        Tmax = 85
        Tmin = 0

        vmax_string = ns_fv * voc_stc * (1 - (coef_voc * (25 - Tmin)))
        vmin_fv_spmp = ns_fv * vop_nmot * (1 - (coef_pmax * (25 - Tmax)))
        vmax_fv_spmp = ns_fv * vop_nmot * (1 - (coef_pmax * (25 - Tmin)))
        iscmax_fv = np_fv * isc_stc * (1 - (coef_isc * (25 - Tmax)))

        if vmax_string < vmax_cc and vmin_fv_spmp > vmin_spmp and vmax_fv_spmp < vmax_spmp and iscmax_fv < imax_spmp:
            # self.mensagem7.place()
            self.mensagem8["foreground"] = 'deep sky blue'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nTodos os parâmetros estão dentro da faixa do inversor, configuração adequada!"""
            # self.mensagem8.place_forget()

        elif vmax_string > vmax_cc and vmin_fv_spmp > vmin_spmp and vmax_fv_spmp < vmax_spmp and iscmax_fv < imax_spmp:
            # self.mensagem8.place()
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nMáxima tensão CC: {vmax_string:.2f} excede o limite do inversor, redimensione a string!"""
            # self.mensagem7.place_forget()

        elif vmin_fv_spmp < vmin_spmp and vmax_string < vmax_cc and vmax_fv_spmp < vmax_spmp and iscmax_fv < imax_spmp:
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nParâmetro tensão mín.: {vmin_spmp:.2f} abaixo do limite do inversor, redimensione a string!"""


        elif vmax_fv_spmp > vmax_spmp and vmax_string < vmax_cc and vmin_fv_spmp > vmin_spmp and iscmax_fv < imax_spmp:
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nParâmetro Tensão máx.: {vmax_fv_spmp:.2f} acima do limite do inversor, redimensione a string!"""


        elif iscmax_fv > imax_spmp and vmax_string < vmax_cc and vmin_fv_spmp > vmin_spmp and vmax_fv_spmp < vmax_spmp:
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nParâmetro corrente máx.: {iscmax_fv:.2f} acima do limite do inversor, redimensione a string!"""






        elif vmax_string > vmax_cc and vmin_fv_spmp > vmin_spmp and vmax_fv_spmp > vmax_spmp and iscmax_fv < imax_spmp:
            # self.mensagem8.place()
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nMáxima tensão CC: {vmax_string:.2f} excede o limite do inversor, redimensione a string!\nParâmetro Tensão máx.: {vmax_fv_spmp:.2f} acima do limite do inversor, redimensione a string!"""


        elif vmax_string > vmax_cc and vmin_fv_spmp > vmin_spmp and vmax_fv_spmp < vmax_spmp and iscmax_fv > imax_spmp:
            # self.mensagem8.place()
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nMáxima tensão CC: {vmax_string:.2f} excede o limite do inversor, redimensione a string!\nParâmetro corrente máx.: {iscmax_fv:.2f} acima do limite do inversor, redimensione a string!"""


        elif vmax_string > vmax_cc and vmin_fv_spmp > vmin_spmp and vmax_fv_spmp > vmax_spmp and iscmax_fv > imax_spmp:
            # self.mensagem8.place()
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nMáxima tensão CC: {vmax_string:.2f} excede o limite do inversor, redimensione a string!\nParâmetro Tensão máx.: {vmax_fv_spmp:.2f} acima do limite do inversor, redimensione a string!\nParâmetro corrente máx.: {iscmax_fv:.2f} acima do limite do inversor, redimensione a string!"""


        elif vmin_fv_spmp < vmin_spmp and vmax_string < vmax_cc and vmax_fv_spmp < vmax_spmp and iscmax_fv > imax_spmp:
            self.mensagem8["foreground"] = 'red'
            self.mensagem8["font"] = ("Arial", "11", "bold")
            self.mensagem8["text"] = f"""{vmax_string:.2f} < {vmax_cc:.2f}\n{vmin_fv_spmp:.2f} > {vmin_spmp:.2f}\n{vmax_fv_spmp:.2f} < {vmax_spmp:.2f}\n{iscmax_fv:.2f} < {imax_spmp:.2f}\n\nParâmetro tensão mín.: {vmin_spmp:.2f} abaixo do limite do inversor, redimensione a string!\nParâmetro corrente máx.: {iscmax_fv:.2f} acima do limite do inversor, redimensione a string!"""

            # self.mensagem7["font"] = ("Arial", "10", "bold")
            # self.mensagem7["text"] = f"""Tensão de {vmax_string} excede o limite do inversor, reduza o n° de módulos em série e tente novamente!"""








    def Janela_localizacao(self):

        self.Janela_localizacao = Toplevel()
        self.Janela_localizacao.geometry("760x300")
        self.Janela_localizacao.minsize(760, 300)
        self.Janela_localizacao.maxsize(760, 300)
        self.Janela_localizacao.configure(bg='gray18')
        self.Janela_localizacao.title("Localização e situação")
        self.Janela_localizacao.iconbitmap('baixados.ico')

        self.fontePadrao = ("Arial", "10")

        self.titulo = Label(self.Janela_localizacao, text="Localização e situação", bg='gray18', fg='orange')
        self.titulo["font"] = ("Arial", "12", "bold")
        self.titulo.place(x=300, y=20)

        self.endereco1Label = Label(self.Janela_localizacao, text="Insira o endereço:", font=self.fontePadrao, bg='gray18', fg='white')
        self.endereco1Label.place(x=60, y=110)

        self.endereco1 = Entry(self.Janela_localizacao)
        self.endereco1["width"] = 60
        self.endereco1["font"] = self.fontePadrao
        self.endereco1.place(x=207, y=110)

        self.clienteLabel = Label(self.Janela_localizacao, text="Insira o nome do cliente:", font=self.fontePadrao, bg='gray18', fg='white')
        self.clienteLabel.place(x=60, y=140)

        self.cliente = Entry(self.Janela_localizacao)
        self.cliente["width"] = 60
        self.cliente["font"] = self.fontePadrao
        self.cliente.place(x=207, y=140)

        self.calcular9 = Button(self.Janela_localizacao, bg='white', fg='darkorange2')
        self.calcular9["text"] = "Gerar Localização"
        self.calcular9["font"] = ("Calibri", "11", "bold")
        self.calcular9["width"] = 30
        self.calcular9["command"] = self.Localizacao
        self.calcular9.place(x=260, y=200)

    def Localizacao(self):

        end = self.endereco1.get()
        client = self.cliente.get().upper()

        # Iniciar Webdriver
        options = Options()
        #options.add_argument("--headless")
        options.add_argument("--handless")
        driver = webdriver.Chrome(chrome_options = options)

        # Acessar endereço
        #driver.get("https://www.google.com.br/maps/preview")
        driver.get("https://www.google.com.br/maps/@-29.9690705,-51.1062926,4006m/data=!3m1!1e3")

        # Esperar até a caixa de busca aparecer
        searchField = WebDriverWait(driver, 10).until(
          EC.presence_of_element_located((By.CSS_SELECTOR, "[name=q]"))
        )

        # Alterar o tamanho da tela
        driver.set_window_size(700, 850)

        # Digitar a "enderço" na caixa de busca e digitar <Enter>

        searchField.send_keys(end)
        searchField.send_keys(Keys.ENTER)

        # Esperar por 5 segundos
        time.sleep(5)

        # Atualiza a pagina
        driver.refresh()

        # Esperar por 5 segundos
        time.sleep(5)

        searchField1 = WebDriverWait(driver, 10).until(
          EC.presence_of_element_located((By.CSS_SELECTOR, "[class=EIbCs]"))
        )

        searchField1.click()

        time.sleep(5)

        driver.save_screenshot(f"{client}.png")

        driver.quit()





    def Janela_monit_remoto(self):

        wb_obj = op.load_workbook('LISTA DE CLIENTES (MONITORAMENTO REMOTO).xlsx')
        sheet_obj = wb_obj.active

        self.Janela_monit_remoto = Toplevel()
        self.Janela_monit_remoto.geometry("950x450")
        self.Janela_monit_remoto.minsize(950, 450)
        self.Janela_monit_remoto.maxsize(950, 450)
        self.Janela_monit_remoto.configure(bg='gray18')
        self.Janela_monit_remoto.title("Monitoramento Remoto")
        self.Janela_monit_remoto.iconbitmap('baixados.ico')

        self.fontePadrao = ("Arial", "10")

        self.titulo = Label(self.Janela_monit_remoto, text="Cadastro Cliente - Monitoramento Remoto", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "12", "bold")
        self.titulo.place(x=350, y=20)

        self.clienteLabel = Label(self.Janela_monit_remoto, text="Insira o nome:", font=self.fontePadrao, fg='white', bg='gray18')
        self.clienteLabel.place(x=220, y=80)

        self.cliente = Entry(self.Janela_monit_remoto)
        self.cliente["width"] = 60
        self.cliente["font"] = self.fontePadrao
        self.cliente.place(x=330, y=80)

        self.fabricanteLabel = Label(self.Janela_monit_remoto, text="Insira o fabricante:", font=self.fontePadrao, fg='white', bg='gray18')
        self.fabricanteLabel.place(x=220, y=110)

        self.fabricante = Entry(self.Janela_monit_remoto)
        self.fabricante["width"] = 60
        self.fabricante["font"] = self.fontePadrao
        self.fabricante.place(x=330, y=110)

        self.loginLabel = Label(self.Janela_monit_remoto, text="Insira o login:", font=self.fontePadrao, fg='white', bg='gray18')
        self.loginLabel.place(x=220, y=140)

        self.login = Entry(self.Janela_monit_remoto)
        self.login["width"] = 60
        self.login["font"] = self.fontePadrao
        self.login.place(x=330, y=140)

        self.senhaLabel = Label(self.Janela_monit_remoto, text="Insira a senha:", font=self.fontePadrao, fg='white', bg='gray18')
        self.senhaLabel.place(x=220, y=170)

        self.senha = Entry(self.Janela_monit_remoto)
        self.senha["width"] = 60
        self.senha["font"] = self.fontePadrao
        self.senha.place(x=330, y=170)

        self.calcular10 = Button(self.Janela_monit_remoto, fg='darkorange2', bg='white')
        self.calcular10["text"] = "Cadastrar cliente"
        self.calcular10["font"] = ("Calibri", "11", "bold")
        self.calcular10["width"] = 30
        self.calcular10["command"] = self.Cadastro_monit_remoto
        self.calcular10.place(x=360, y=210)


        self.lista_cliente = Label(self.Janela_monit_remoto, text="Clientes cadastrados:", font=("Arial", 10, "bold", "italic"), fg='white', bg='gray18')
        self.lista_cliente.place(x=290, y=260)
        ##coluna = bd['Modelo']
        lista = []
        ##inversor = bd['Modelo']

        max_row = sheet_obj.max_row
        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=1)
            lista.append(cell_obj.value)

        self.variable_cliente = tk.StringVar(self.Janela_monit_remoto)
        self.variable_cliente.set(lista[0])

        self.lista = ttk.Combobox(self.Janela_monit_remoto, textvariable=self.variable_cliente)
        self.lista["width"] = 30
        self.lista.place(x=425, y=260)

        self.lista['values'] = [lista[m] for m in range(len(lista))]

        self.calcular10 = Button(self.Janela_monit_remoto, fg='darkorange2', bg='white')
        self.calcular10["text"] = "Consultar Dados"
        self.calcular10["font"] = ("Calibri", "11", "bold")
        self.calcular10["width"] = 30
        self.calcular10["command"] = self.Consulta_monit_remoto
        self.calcular10.place(x=360, y=300)

        self.mensagem9 = Label(self.Janela_monit_remoto, text="", font=self.fontePadrao, foreground='deep sky blue', bg='gray18')
        self.mensagem9.place(x=300, y=370)

    def Cadastro_monit_remoto(self):


       # GERA LISTA DE MÓDULOS

        wb_obj = op.load_workbook('LISTA DE CLIENTES (MONITORAMENTO REMOTO).xlsx')
        sheet_obj = wb_obj.active

        cliente = self.cliente.get()
        fabricante = self.fabricante.get()
        login = self.login.get()
        senha = self.senha.get()


        to_append = [cliente, fabricante, login, senha]

        sheet_obj.append(to_append)

        wb_obj.save('LISTA DE CLIENTES (MONITORAMENTO REMOTO).xlsx')

    def Consulta_monit_remoto(self):


        wb_obj = op.load_workbook('LISTA DE CLIENTES (MONITORAMENTO REMOTO).xlsx')
        sheet_obj = wb_obj.active

        cliente = self.cliente.get()
        login = self.login.get()
        senha = self.senha.get()

        list = int(self.lista.current())

        dados3 = []
        max_col2 = sheet_obj.max_column
        max_row2 = sheet_obj.max_row
        for m in range(1, max_col2 + 1):
            for n in range(1, max_row2 + 1):
                cell_obj = sheet_obj.cell(row=n, column=m)
                dados3.append(cell_obj.value)
                data2 = []
                for o in dados3[list::max_row2]:
                    data2.append(o)

        nome_cliente = data2[0]
        fabricante_cliente = data2[1]
        login_cliente = data2[2]
        senha_cliente = data2[3]

        self.mensagem9["font"] = ("Arial", "11", "bold")
        self.mensagem9["text"] = f"Cliente: {nome_cliente} | Fabricante: {fabricante_cliente}\n Login: {login_cliente} | Senha: {senha_cliente}"




    def Janela_projecao(self):

        # CALCULA A PROJEÇÃO DO MODULO INCLINADO NO PLANO HORIZONTAL

        self.Janela_projecao = Toplevel()
        self.Janela_projecao.geometry("850x450")
        self.Janela_projecao.minsize(850, 450)
        self.Janela_projecao.maxsize(850, 450)
        self.Janela_projecao.configure(bg='gray18')
        self.Janela_projecao.title("Projeção no plano horizontal / Distância entre fileiras de módulos")
        self.Janela_projecao.iconbitmap('baixados.ico')

        wb_obj = op.load_workbook('dados_mod.xlsx')
        sheet_obj = wb_obj.active

        # PROJEÇÃO

        self.titulo = Label(self.Janela_projecao,text="Projeção no plano horizontal / Distância entre fileiras de módulos", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "12", "bold")
        self.titulo.place(x=200, y=20)

        self.modulo = Label(self.Janela_projecao, text="Escolha o módulo:", font=("Arial", 10, "bold", "italic"), fg='white', bg='gray18')
        self.modulo.place(x=300, y=80)

        modulo = []

        max_row = sheet_obj.max_row
        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            modulo.append(cell_obj.value)

        self.variable_mod = tk.StringVar(self.Janela_projecao)
        self.variable_mod.set(modulo[0])

        self.modulo = ttk.Combobox(self.Janela_projecao, textvariable=self.variable_mod)
        self.modulo["width"] = 20
        self.modulo.place(x=445, y=80)

        self.modulo['values'] = [modulo[m] for m in range(len(modulo))]

        self.orientacao = Label(self.Janela_projecao, text="Escolha a orientação:", font=("Arial", 10, "bold", "italic"), fg='white', bg='gray18')
        self.orientacao.place(x=300, y=110)

        orient = ['Selecione', 'Retrato', 'Paisagem']

        self.variable_orientacao = tk.StringVar(self.Janela_projecao)
        self.variable_orientacao.set(orient[0])

        self.orientacao = ttk.Combobox(self.Janela_projecao, textvariable=self.variable_orientacao)
        self.orientacao["width"] = 20
        self.orientacao.place(x=445, y=110)

        self.orientacao['values'] = [orient[m] for m in range(0, 3)]

        self.inclinacaoLabel = Label(self.Janela_projecao, text="Inclinação do módulo:", font=self.fontePadrao, fg='white', bg='gray18')
        self.inclinacaoLabel.place(x=300, y=140)

        self.inclinacao = Entry(self.Janela_projecao)
        self.inclinacao["width"] = 10
        self.inclinacao["font"] = self.fontePadrao
        self.inclinacao.place(x=445, y=140)

        self.fileiraLabel = Label(self.Janela_projecao, text="N° de módulos na fileira:", font=self.fontePadrao, fg='white', bg='gray18')
        self.fileiraLabel.place(x=300, y=170)

        self.fileira = Entry(self.Janela_projecao)
        self.fileira["width"] = 10
        self.fileira["font"] = self.fontePadrao
        self.fileira.place(x=445, y=170)

        self.alturaLabel = Label(self.Janela_projecao, text="Altura do plano horizontal até a base da fileira em m:", font=self.fontePadrao, fg='white', bg='gray18')
        self.alturaLabel.place(x=300, y=200)

        self.altura = Entry(self.Janela_projecao)
        self.altura["width"] = 10
        self.altura["font"] = self.fontePadrao
        self.altura.place(x=610, y=200)

        self.latitudeLabel = Label(self.Janela_projecao, text="Latitude local:",font=self.fontePadrao, fg='white', bg='gray18')
        self.latitudeLabel.place(x=300, y=230)

        self.latitude = Entry(self.Janela_projecao)
        self.latitude["width"] = 10
        self.latitude["font"] = self.fontePadrao
        self.latitude.place(x=445, y=230)

        self.calcular10 = Button(self.Janela_projecao, fg='darkorange2', bg='white')
        self.calcular10["text"] = "Calcular"
        self.calcular10["font"] = ("Calibri", "11", "bold")
        self.calcular10["width"] = 30
        self.calcular10["command"] = self.Calculo_projecao
        self.calcular10.place(x=300, y=270)

        self.mensagem10 = Label(self.Janela_projecao, text="", font=self.fontePadrao, foreground='deep sky blue', bg='gray18')
        self.mensagem10.place(x=200, y=320)

    def Calculo_projecao(self):

        wb_obj = op.load_workbook('dados_mod.xlsx')
        sheet_obj = wb_obj.active

        inc = float(self.inclinacao.get().replace(",", "."))
        n_mod_fileira = int(self.fileira.get().replace(",", "."))
        alt = float(self.altura.get().replace(",", "."))
        lat = float(self.latitude.get().replace(",", "."))
        ori = self.variable_orientacao.get().upper()

        ind = int(self.modulo.current())

        dados4 = []
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        for j in range(3, max_col + 1):
            for k in range(2, max_row + 1):
                cell_obj = sheet_obj.cell(row=k, column=j)
                dados4.append(cell_obj.value)
                data3 = []
                for l in dados4[ind - 1::max_row - 1]:
                    data3.append(l)

        if ori == 'RETRATO':
            comprimento = data3[8]
            largura = data3[9]
        elif ori == 'PAISAGEM':
            comprimento = data3[9]
            largura = data3[8]

        inc_graus = (inc * 3.14159) / 180
        lat_graus = (lat * 3.14159) / 180
        #inc_cos = math.cos(inc_cos_graus)
        #inc_cos_graus = math.degrees(inc_cos_radianos)

        #inc_sin_graus = np.rad2deg(inc)
        #inc_sin_radianos = math.sin(math.degrees(inc))
        #inc_sin_graus = math.degrees(inc_sin_radianos)

        #inc_tan_graus = np.rad2deg(lat)

        proj_horizontal = math.cos(inc_graus) * ((comprimento / 1000) * n_mod_fileira)

        altura = (math.sin(inc_graus) * ((comprimento / 1000) * n_mod_fileira)) + alt

        distancia_fileira = altura / math.tan(((61 * 3.14159) / 180) - lat_graus)

        self.mensagem10["font"] = ("Arial", "11", "bold")
        self.mensagem10["text"] = f"A projeção do módulo no plano horizontal é de {proj_horizontal:.2f} m \nA altura da fileira de módulos é de {altura:.2f} m\n A distância entre as fileiras de módulos deverá ser de {distancia_fileira:.2f} m"






    def Janela_automacao(self):

        self.Janela_automacao = Toplevel()
        self.Janela_automacao.geometry("830x450")
        self.Janela_automacao.minsize(830, 450)
        self.Janela_automacao.maxsize(830, 450)
        self.Janela_automacao.configure(bg='gray18')
        self.Janela_automacao.title("Automação de diagrama unifilar")
        self.Janela_automacao.iconbitmap('baixados.ico')

        self.fontePadrao = ("Arial", "10")

        self.titulo = Label(self.Janela_automacao, text="Automação de diagrama unifilar", fg='orange', bg='gray18')
        self.titulo["font"] = ("Arial", "12", "bold")
        self.titulo.place(x=320, y=20)

        wb_obj = op.load_workbook('dados_mod.xlsx')
        sheet_obj = wb_obj.active

        wb_obj1 = op.load_workbook('dados_inv.xlsx')
        sheet_obj1 = wb_obj1.active

        self.modulo = Label(self.Janela_automacao, text="Escolha o módulo:", font=("Arial", 10, "bold", "italic"), fg='white', bg='gray18')
        self.modulo.place(x=50, y=80)
        ##coluna = bd['Modelo']
        modulo = []
        ##inversor = bd['Modelo']

        max_row = sheet_obj.max_row
        for i in range(1, max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            modulo.append(cell_obj.value)

        self.variable_mod = tk.StringVar(self.Janela_automacao)
        self.variable_mod.set(modulo[0])

        self.modulo = ttk.Combobox(self.Janela_automacao, textvariable=self.variable_mod)
        self.modulo["width"] = 20
        self.modulo.place(x=238, y=80)

        self.modulo['values'] = [modulo[m] for m in range(len(modulo))]

        self.inversor = Label(self.Janela_automacao, text="Escolha o inversor:", font=("Arial", 10, "bold", "italic"), fg='white', bg='gray18')
        self.inversor.place(x=400, y=80)
        ##coluna = bd['Modelo']
        inversor = []
        ##inversor = bd['Modelo']

        max_row1 = sheet_obj1.max_row
        for i in range(1, max_row1 + 1):
            cell_obj1 = sheet_obj1.cell(row=i, column=2)
            inversor.append(cell_obj1.value)

        self.variable_inv = tk.StringVar(self.Janela_automacao)
        self.variable_inv.set(inversor[0])

        self.inversor = ttk.Combobox(self.Janela_automacao, textvariable=self.variable_inv)
        self.inversor["width"] = 20
        self.inversor.place(x=555, y=80)

        self.inversor['values'] = [inversor[m] for m in range(len(inversor))]

        self.potencia_usinaLabel = Label(self.Janela_automacao, text="Insira a potência da usina (kW):",font=self.fontePadrao, fg='white', bg='gray18')
        self.potencia_usinaLabel.place(x=50, y=110)

        self.potencia_usina = Entry(self.Janela_automacao)
        self.potencia_usina["width"] = 35
        self.potencia_usina["font"] = self.fontePadrao
        self.potencia_usina.place(x=238, y=110)

        self.condutorLabel = Label(self.Janela_automacao, text="Insira os condutores:", font=self.fontePadrao, fg='white', bg='gray18')
        self.condutorLabel.place(x=50, y=140)

        self.condutor = Entry(self.Janela_automacao)
        self.condutor["width"] = 35
        self.condutor["font"] = self.fontePadrao
        self.condutor.place(x=238, y=140)

        self.disjuntorLabel = Label(self.Janela_automacao, text="Insira o disjuntor:", font=self.fontePadrao, fg='white', bg='gray18')
        self.disjuntorLabel.place(x=50, y=170)

        self.disjuntor = Entry(self.Janela_automacao)
        self.disjuntor["width"] = 20
        self.disjuntor["font"] = self.fontePadrao
        self.disjuntor.place(x=238, y=170)

        self.disjuntor_geralLabel = Label(self.Janela_automacao, text="Insira o disjuntor geral:",font=self.fontePadrao, fg='white', bg='gray18')
        self.disjuntor_geralLabel.place(x=50, y=200)

        self.disjuntor_geral = Entry(self.Janela_automacao)
        self.disjuntor_geral["width"] = 20
        self.disjuntor_geral["font"] = self.fontePadrao
        self.disjuntor_geral.place(x=238, y=200)

        self.disjuntor_medicaoLabel = Label(self.Janela_automacao, text="Insira o disjuntor de medição:",font=self.fontePadrao, fg='white', bg='gray18')
        self.disjuntor_medicaoLabel.place(x=50, y=230)

        self.disjuntor_medicao = Entry(self.Janela_automacao)
        self.disjuntor_medicao["width"] = 20
        self.disjuntor_medicao["font"] = self.fontePadrao
        self.disjuntor_medicao.place(x=238, y=230)

        self.concessionariaLabel = Label(self.Janela_automacao, text="Insira a concessionária:", font=self.fontePadrao, fg='white', bg='gray18')
        self.concessionariaLabel.place(x=50, y=260)

        self.concessionaria = Entry(self.Janela_automacao)
        self.concessionaria["width"] = 20
        self.concessionaria["font"] = self.fontePadrao
        self.concessionaria.place(x=238, y=260)

        self.enderecoLabel = Label(self.Janela_automacao, text="Insira o endereço:", font=self.fontePadrao, fg='white', bg='gray18')
        self.enderecoLabel.place(x=50, y=290)

        self.endereco = Entry(self.Janela_automacao)
        self.endereco["width"] = 71
        self.endereco["font"] = self.fontePadrao
        self.endereco.place(x=238, y=290)

        self.proprietarioLabel = Label(self.Janela_automacao, text="Proprietário:", font=self.fontePadrao, fg='white', bg='gray18')
        self.proprietarioLabel.place(x=400, y=170)

        self.proprietario = Entry(self.Janela_automacao)
        self.proprietario["width"] = 25
        self.proprietario["font"] = self.fontePadrao
        self.proprietario.place(x=555, y=170)

        self.projetoLabel = Label(self.Janela_automacao, text="Insira o titulo do projeto:", font=self.fontePadrao, fg='white', bg='gray18')
        self.projetoLabel.place(x=400, y=200)

        self.projeto = Entry(self.Janela_automacao)
        self.projeto["width"] = 25
        self.projeto["font"] = self.fontePadrao
        self.projeto.place(x=555, y=200)

        self.dataLabel = Label(self.Janela_automacao, text="Insira a data:", font=self.fontePadrao, fg='white', bg='gray18')
        self.dataLabel.place(x=400, y=230)

        self.data = Entry(self.Janela_automacao)
        self.data["width"] = 25
        self.data["font"] = self.fontePadrao
        self.data.place(x=555, y=230)

        self.codigoLabel = Label(self.Janela_automacao, text="Insira o código do projeto:", font=self.fontePadrao, fg='white', bg='gray18')
        self.codigoLabel.place(x=400, y=260)

        self.codigo = Entry(self.Janela_automacao)
        self.codigo["width"] = 25
        self.codigo["font"] = self.fontePadrao
        self.codigo.place(x=555, y=260)

        self.observacaoLabel = Label(self.Janela_automacao, text="Observações:", font=self.fontePadrao, fg='white', bg='gray18')
        self.observacaoLabel.place(x=50, y=320)

        self.observacao = Entry(self.Janela_automacao)
        self.observacao["width"] = 71
        self.observacao["font"] = self.fontePadrao
        self.observacao.place(x=238, y=320)

        self.calcular11 = Button(self.Janela_automacao, fg='darkorange2', bg='white')
        self.calcular11["text"] = "Gerar Unifilar"
        self.calcular11["font"] = ("Calibri", "11", "bold")
        self.calcular11["width"] = 30
        self.calcular11["command"] = self.Unifilar_automacao
        self.calcular11.place(x=310, y=380)

    def Unifilar_automacao(self):

        # Módulo:

        wb_obj = op.load_workbook('dados_mod.xlsx')
        sheet_obj = wb_obj.active

        ind = int(self.modulo.current())

        dados1 = []
        max_col = sheet_obj.max_column
        max_row = sheet_obj.max_row
        for j in range(1, max_col + 1):
            for k in range(2, max_row + 1):
                cell_obj = sheet_obj.cell(row=k, column=j)
                dados1.append(cell_obj.value)
                data = []
                for l in dados1[ind - 1::max_row - 1]:
                    data.append(l)

        # Inversor:

        wb_obj1 = op.load_workbook('dados_inv.xlsx')
        sheet_obj1 = wb_obj1.active

        ind1 = int(self.inversor.current())

        dados2 = []
        max_col1 = sheet_obj1.max_column
        max_row1 = sheet_obj1.max_row
        for m in range(1, max_col1 + 1):
            for n in range(2, max_row1 + 1):
                cell_obj1 = sheet_obj1.cell(row=n, column=m)
                dados2.append(cell_obj1.value)
                data1 = []
                for o in dados2[ind1 - 1::max_row1 - 1]:
                    data1.append(o)

        pot_usina = float(self.potencia_usina.get().replace(",", "."))
        pot_modulos = data[2]
        n_modulos = int((pot_usina * 1000) / pot_modulos)
        fabric = data[0]
        model = data[1]

        fabric_inv = data1[0]
        model_inv = data1[1]
        pot_inv = data1[2] / 1000

        condut = self.condutor.get()
        disj = self.disjuntor.get().upper()
        disj_geral = self.disjuntor_geral.get().upper()
        disj_medicao = self.disjuntor_medicao.get().upper()
        conce = self.concessionaria.get().upper()
        endereco = self.endereco.get().upper()
        proprietario = self.proprietario.get().upper()
        proj = self.projeto.get().upper()
        data = self.data.get().upper()
        cod = self.codigo.get().upper()
        obs = self.observacao.get().upper()

        pyautogui.moveTo(1060, 1073)
        time.sleep(1)
        pyautogui.hotkey('winleft', 'e')
        time.sleep(2)
        pyautogui.hotkey('winleft', 'up')
        time.sleep(2)
        pyautogui.click(515, 155)
        time.sleep(1)
        pyautogui.write("G:/Douglas/UNIFILAR AUTOMATIZADO.dwg")
        time.sleep(1)
        pyautogui.press("enter")
        time.sleep(9)
        pyautogui.moveTo(1182, 691)
        pyautogui.doubleClick(1182, 691)
        pyautogui.write(f"GERADOR FOTOVOLTAICO {pot_usina} kW")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(1144, 701)
        pyautogui.click(1144, 701)
        pyautogui.write(f"{n_modulos} MODULOS {pot_modulos} Wp")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(1186, 711)
        pyautogui.click(1186, 711)
        pyautogui.write(f"{fabric} - {model}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(965, 686)
        pyautogui.click(965, 686)
        pyautogui.write(f"INVERSOR {fabric_inv} - {model_inv}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(944, 696)
        pyautogui.click(944, 696)
        pyautogui.write(f"POTENCIA {pot_inv} kW")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(852, 714)
        pyautogui.click(852, 714)
        pyautogui.write(f"{condut}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(800, 754)
        pyautogui.click(800, 754)
        pyautogui.write(f"DISJ. {disj} A")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(785, 651)
        pyautogui.click(785, 651)
        pyautogui.write(f"{condut}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(733, 592)
        pyautogui.click(733, 592)
        pyautogui.write(f"DISJ. {disj} A")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(546, 555)
        pyautogui.click(546, 555)
        pyautogui.write(f"DISJ. {disj_geral} A")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(434, 360)
        pyautogui.click(434, 360)
        pyautogui.write(f"{disj_medicao} A")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(402, 189)
        pyautogui.click(402, 189)
        pyautogui.write(f"REDE BT {conce}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(177, 912)
        pyautogui.click(177, 912)

        time.sleep(1)
        pyautogui.moveTo(1067, 772)
        pyautogui.doubleClick(1067, 772)
        pyautogui.write(f"{endereco}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(1035, 808)
        pyautogui.doubleClick(1035, 808)
        pyautogui.write(f"PROPRIETARIO: {proprietario}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(1107, 858)
        pyautogui.doubleClick(1107, 858)
        pyautogui.write(f"{proj}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(1080, 879)
        pyautogui.doubleClick(1080, 879)
        pyautogui.write(f"{data}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(1238, 829)
        pyautogui.doubleClick(1238, 829)
        pyautogui.write(f"{cod}")
        pyautogui.press("enter")

        time.sleep(1)
        pyautogui.moveTo(1069, 680)
        pyautogui.doubleClick(1069, 680)
        pyautogui.write(f"{obs}")
        pyautogui.press("enter")

        time.sleep(2)
        pyautogui.moveTo(180, 31)
        pyautogui.click(180, 31)

        time.sleep(2)
        pyautogui.moveTo(217, 191)
        pyautogui.click(217, 191)
        time.sleep(3)
        pyautogui.write(f"{proprietario[:4]}")

        time.sleep(2)
        pyautogui.moveTo(730, 736)
        pyautogui.click(730, 736)
        time.sleep(2)
        pyautogui.press("enter")

        time.sleep(2)
        pyautogui.moveTo(1130, 764)
        pyautogui.click(1130, 764)

        time.sleep(2)
        pyautogui.moveTo(862, 627)
        pyautogui.click(862, 627)

        time.sleep(2)
        pyautogui.hotkey('ctrl', 'shift', 's')
        time.sleep(1)
        pyautogui.write(f"UNIFILAR {proprietario}")
        time.sleep(1)
        pyautogui.press("enter")

        time.sleep(5)
        pyautogui.hotkey('ctrl', 'p')
        time.sleep(2)
        pyautogui.press("enter")
        time.sleep(2)
        pyautogui.write(f"UNIFILAR {proprietario}")
        time.sleep(1)
        pyautogui.press("enter")


#time.sleep(2)
#position = pyautogui.position()
#print(position)


root = tk.Tk()
Application(root)
#root.overrideredirect(True)
root.title('Projetos FV')
#width = root.winfo_screenwidth()
#height = root.winfo_screenheight()
bg1 = tk.PhotoImage(file="solar1.png")
w = tk.Label(root, image=bg1, bg="gray18")
w.bg1 = bg1
w.place(x=190, y=40)
#canvas1 = Canvas(root, width=760,height=300)
#canvas1.pack(fill="both", expand=True)
#canvas1.create_image(0, 0, image=bg1,anchor="nw")
root.geometry("750x570")
root.minsize(750, 570)
root.maxsize(750, 570)
root.configure(bg='gray18')
root.iconbitmap('baixados.ico')
root.mainloop()